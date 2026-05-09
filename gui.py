"""
Gestionale Officina – Garage Tito
Avvio: python gui.py
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import os, shutil, datetime, threading, subprocess, sys
from db_connection import get_connection

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DIR_DOCS  = os.path.join(BASE_DIR, "documenti")
DIR_FATT  = os.path.join(BASE_DIR, "fatture")
TEMPLATE  = os.path.join(BASE_DIR, "RICEVUTA_TEMPLATE.xlsm")

# ── VERSIONE ──────────────────────────────────────────────────
VERSIONE_CORRENTE = "1.0.0"

# !! CAMBIA QUESTO con il tuo username GitHub e nome repository !!
GITHUB_USER = "gab91205"
GITHUB_REPO = "Gestionale-Officina"
# URL del file version.txt nella release più recente
GITHUB_VERSION_URL = f"https://raw.githubusercontent.com/{GITHUB_USER}/{GITHUB_REPO}/main/version.txt"
GITHUB_GUI_URL     = f"https://raw.githubusercontent.com/{GITHUB_USER}/{GITHUB_REPO}/main/gui.py"

for d in (DIR_DOCS, DIR_FATT):
    os.makedirs(d, exist_ok=True)

# ── DB ────────────────────────────────────────────────────────
try:
    _t = get_connection(); _t.close()
except Exception as e:
    import tkinter as _tk
    _r = _tk.Tk(); _r.withdraw()
    _tk.messagebox.showerror("Errore DB",
        f"Impossibile aprire il database:\n{e}")
    sys.exit(1)

def db_fetch(sql, params=()):
    """Esegue una SELECT e restituisce lista di dict."""
    # Converte placeholder MySQL (%s) → SQLite (?)
    sql = sql.replace("%s", "?")
    conn = get_connection()
    try:
        cur = conn.execute(sql, params)
        cols = [d[0] for d in cur.description] if cur.description else []
        return [dict(zip(cols, row)) for row in cur.fetchall()]
    finally:
        conn.close()

def db_commit(sql, params=()):
    """Esegue INSERT/UPDATE/DELETE e restituisce lastrowid."""
    sql = sql.replace("%s", "?")
    conn = get_connection()
    try:
        cur = conn.execute(sql, params)
        conn.commit()
        return cur.lastrowid
    finally:
        conn.close()

def _auto_migra():
    """
    Crea tutte le tabelle se non esistono e aggiunge colonne mancanti.
    Compatibile SQLite — usa PRAGMA invece di SHOW COLUMNS.
    """
    try:
        conn = get_connection()

        conn.executescript("""
        CREATE TABLE IF NOT EXISTS clienti (
            id_cliente      INTEGER PRIMARY KEY AUTOINCREMENT,
            nome            TEXT,
            cognome         TEXT NOT NULL DEFAULT '',
            telefono        TEXT,
            email           TEXT,
            indirizzo       TEXT,
            cap             TEXT,
            data_creazione  TEXT,
            tipo_cliente    TEXT DEFAULT 'privato',
            ragione_sociale TEXT
        );
        CREATE TABLE IF NOT EXISTS veicoli (
            id_veicolo  INTEGER PRIMARY KEY AUTOINCREMENT,
            id_cliente  INTEGER REFERENCES clienti(id_cliente) ON DELETE CASCADE,
            targa       TEXT,
            marca       TEXT,
            modello     TEXT,
            anno        INTEGER,
            chilometri  INTEGER DEFAULT 0,
            telaio      TEXT,
            carta_grigia_path TEXT
        );
        CREATE TABLE IF NOT EXISTS storico_interventi (
            id_intervento INTEGER PRIMARY KEY AUTOINCREMENT,
            id_veicolo    INTEGER NOT NULL REFERENCES veicoli(id_veicolo) ON DELETE CASCADE,
            data_lavoro   TEXT NOT NULL,
            descrizione   TEXT NOT NULL,
            costo         REAL DEFAULT 0,
            note          TEXT
        );
        CREATE TABLE IF NOT EXISTS fatture (
            id_fattura      INTEGER PRIMARY KEY AUTOINCREMENT,
            id_cliente      INTEGER REFERENCES clienti(id_cliente),
            id_veicolo      INTEGER REFERENCES veicoli(id_veicolo),
            data_fattura    TEXT,
            totale          REAL,
            pagata          INTEGER DEFAULT 0,
            file_path       TEXT,
            id_storico      INTEGER,
            numero_fattura  TEXT,
            iva_applicata   INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS voci_fattura (
            id_voce     INTEGER PRIMARY KEY AUTOINCREMENT,
            id_fattura  INTEGER NOT NULL REFERENCES fatture(id_fattura) ON DELETE CASCADE,
            descrizione TEXT,
            quantita    REAL DEFAULT 1,
            prezzo      REAL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS magazzino_gomme (
            id_gomma        INTEGER PRIMARY KEY AUTOINCREMENT,
            id_cliente      INTEGER REFERENCES clienti(id_cliente) ON DELETE SET NULL,
            targa_veicolo   TEXT,
            marca           TEXT NOT NULL,
            modello         TEXT,
            stagione        TEXT DEFAULT 'Estiva',
            misura          TEXT,
            quantita        INTEGER DEFAULT 1,
            stato           TEXT DEFAULT 'Nuovo',
            posizione       TEXT,
            deposito_pagato INTEGER DEFAULT 0,
            montate         INTEGER DEFAULT 0,
            note            TEXT
        );
        CREATE TABLE IF NOT EXISTS numero_fattura_config (
            id              INTEGER PRIMARY KEY DEFAULT 1,
            prossimo_numero INTEGER
        );
        """)

        # INSERT OR IGNORE per il record di config
        conn.execute("INSERT OR IGNORE INTO numero_fattura_config (id) VALUES (1)")
        conn.commit()
        conn.close()
    except Exception:
        pass

_auto_migra()

def apri_file(path):
    if sys.platform == "win32":
        os.startfile(path)
    elif sys.platform == "darwin":
        subprocess.call(["open", path])

IVA = 0.081   # IVA svizzera 8.1%

def _riavvia_watcher_tutte_fatture():
    """
    All'avvio del gestionale, rilancia il watcher per tutte le fatture
    che hanno un file Excel e uno storico collegato.
    Così le modifiche fatte il giorno dopo vengono comunque monitorate.
    """
    try:
        rows = db_fetch(
            "SELECT f.id_fattura, f.file_path, f.id_storico, f.id_veicolo, f.totale "
            "FROM fatture f "
            "WHERE f.file_path IS NOT NULL AND f.id_storico IS NOT NULL")
        for r in rows:
            path = os.path.join(BASE_DIR, r["file_path"]) if r["file_path"] else None
            if path and os.path.exists(path):
                _avvia_watcher_fattura(
                    path,
                    r["id_storico"],
                    r["id_veicolo"],
                    float(r["totale"] or 0))
    except Exception:
        pass  # Non bloccare l'avvio

def _avvia_watcher_fattura(excel_path, id_storico, id_veicolo, totale_originale):
    """
    Monitora il file Excel. Quando viene salvato legge H46 direttamente
    (totale finale CHF con IVA già inclusa dal template Excel).
    NON fa nessun calcolo — usa solo il valore che Excel ha già calcolato.
    """
    import time

    def _leggi_H44_com(path):
        """Legge H46 dall'Excel già aperto o aprendolo silenziosamente."""
        nome_file = os.path.basename(path)
        # Prova Excel già aperto
        try:
            import win32com.client
            xl = win32com.client.GetActiveObject("Excel.Application")
            for wb in xl.Workbooks:
                if os.path.basename(wb.FullName).lower() == nome_file.lower():
                    ws = wb.Sheets("Ricevuta")
                    return _estrai_com(ws)
        except Exception:
            pass
        # Apri silenziosamente
        try:
            import win32com.client
            xl = win32com.client.Dispatch("Excel.Application")
            xl.Visible = False
            xl.DisplayAlerts = False
            wb = xl.Workbooks.Open(path, ReadOnly=True, UpdateLinks=False)
            wb.Application.CalculateFull()
            ws = wb.Sheets("Ricevuta")
            result = _estrai_com(ws)
            wb.Close(False)
            xl.Quit()
            return result
        except Exception:
            pass
        return None, None, None, None

    def _estrai_com(ws):
        """Estrae i dati dalla worksheet COM — legge H46 senza calcoli."""
        num_b7   = ws.Range("B7").Value
        desc_b17 = str(ws.Range("B16").Value or "")
        # H46 = totale CHF finale (già calcolato da Excel con IVA)
        totale   = ws.Range("H46").Value
        righe = []
        for row in range(18, 44):
            desc = ws.Range(f"B{row}").Value
            qty  = ws.Range(f"E{row}").Value
            prc  = ws.Range(f"F{row}").Value
            if desc:
                try:
                    q, p = float(qty or 0), float(prc or 0)
                    righe.append(f"- {desc}  (Qt {q:.0f} x CHF {p:.2f})" if q and p else f"- {desc}")
                except Exception:
                    righe.append(f"- {desc}")
        try:
            totale = round(float(totale), 2) if totale else None
        except Exception:
            totale = None
        return num_b7, desc_b17, righe, totale

    def _leggi_H44_openpyxl(path):
        """Fallback openpyxl — legge H46 cached."""
        try:
            from openpyxl import load_workbook
            wb2 = load_workbook(path, data_only=True)
            ws2 = wb2["Ricevuta"]
            num_b7   = ws2["B7"].value
            desc_b17 = str(ws2["B16"].value or "")
            totale   = ws2["H46"].value
            righe = []
            for row in range(18, 44):
                desc = ws2[f"B{row}"].value
                qty  = ws2[f"E{row}"].value
                prc  = ws2[f"F{row}"].value
                if desc:
                    try:
                        q, p = float(qty or 0), float(prc or 0)
                        righe.append(f"- {desc}  (Qt {q:.0f} x CHF {p:.2f})" if q and p else f"- {desc}")
                    except Exception:
                        righe.append(f"- {desc}")
            try:
                totale = round(float(totale), 2) if totale else None
            except Exception:
                totale = None
            return num_b7, desc_b17, righe, totale
        except Exception:
            return None, None, None, None

    def _watch():
        try:
            mtime_prec = os.path.getmtime(excel_path)
        except Exception:
            return

        while True:
            time.sleep(5)
            try:
                mtime_nuovo = os.path.getmtime(excel_path)
            except Exception:
                break

            if mtime_nuovo != mtime_prec:
                mtime_prec = mtime_nuovo
                time.sleep(2)

                num_b7, desc_b17, righe, totale = _leggi_H44_com(excel_path)
                if righe is None:
                    num_b7, desc_b17, righe, totale = _leggi_H44_openpyxl(excel_path)
                if righe is None:
                    continue

                testo = (desc_b17 + "\n") if desc_b17 else ""
                testo += "\n".join(righe) if righe else "Fattura senza voci"

                # Usa H45 (totale finale con IVA calcolato da Excel)
                # Se None (formula non cached), somma manualmente le voci * IVA
                totale_finale = totale if totale and totale > 0 else None
                if not totale_finale:
                    try:
                        from openpyxl import load_workbook as _lw
                        _wb = _lw(excel_path, data_only=True)
                        _ws = _wb["Ricevuta"]
                        imp = sum(
                            float(_ws[f"E{r}"].value or 0) * float(_ws[f"F{r}"].value or 0)
                            for r in range(18, 44) if _ws[f"B{r}"].value)
                        totale_finale = round(imp * (1 + IVA), 2) if imp > 0 else totale_originale
                    except Exception:
                        totale_finale = totale_originale

                try:
                    db_commit(
                        "UPDATE storico_interventi SET descrizione=%s, costo=%s "
                        "WHERE id_intervento=%s",
                        (testo, totale_finale, id_storico))
                    rows_fat = db_fetch(
                        "SELECT id_fattura FROM fatture WHERE id_storico=%s", (id_storico,))
                    if rows_fat:
                        fid = rows_fat[0]["id_fattura"]
                        db_commit(
                            "UPDATE fatture SET totale=%s, iva_applicata=TRUE "
                            "WHERE id_fattura=%s",
                            (totale_finale, fid))
                        if num_b7:
                            db_commit(
                                "UPDATE fatture SET numero_fattura=%s WHERE id_fattura=%s",
                                (str(num_b7), fid))
                except Exception as ex:
                    try:
                        with open(os.path.join(BASE_DIR, "watcher_error.log"), "a") as lf:
                            lf.write(f"{datetime.datetime.now()} | {ex}\n")
                    except Exception:
                        pass

    t = threading.Thread(target=_watch, daemon=True)
    t.start()

# ── NUMERO FATTURA formato AAMMNN (es. 260401) ───────────────
def _excel_date(d):
    return (d - __import__("datetime").date(1899, 12, 30)).days

# ── GESTIONE NUMERO FATTURA ───────────────────────────────────
# Formato AAMMNN: AA=anno, MM=mese, NN=progressivo mensile
# Es: 260401, 260402, ... 260501 (mese dopo)
# Fonte di verità: DB (fatture) + numero_fattura_config (override manuale)

def _calc_prossimo_da_db():
    """
    Formato numero: AA + NNNN  (es. 260201, 260202, ..., 260299, 260301...)
    - Prende tutti i numeri esistenti nel DB
    - Se c'è un buco (fattura cancellata) usa quello
    - Altrimenti usa ultimo+1
    - Se DB vuoto parte da AA0201
    """
    rows = db_fetch(
        "SELECT CAST(numero_fattura AS UNSIGNED) AS n FROM fatture "
        "WHERE numero_fattura IS NOT NULL AND numero_fattura != '' "
        "ORDER BY n")

    if not rows:
        aa = datetime.date.today().strftime("%y")
        return int(f"{aa}0201")

    numeri = sorted([int(r["n"]) for r in rows if r["n"]])

    # Cerca buchi: se 260202 e 260204 esistono, 260203 è libero
    for i in range(len(numeri) - 1):
        if numeri[i+1] - numeri[i] > 1:
            return numeri[i] + 1

    # Nessun buco: prossimo dopo l'ultimo
    return numeri[-1] + 1

def _leggi_numero_config():
    """Legge il numero manuale dal config senza consumarlo. None se non impostato."""
    try:
        cfg = db_fetch("SELECT prossimo_numero FROM numero_fattura_config WHERE id=1")
        if cfg and cfg[0]["prossimo_numero"]:
            return int(cfg[0]["prossimo_numero"])
    except Exception:
        pass
    return None

def prossimo_numero_fattura():
    """
    Restituisce il prossimo numero fattura.
    Se c'è un override manuale usa quello e lo cancella (usa una sola volta).
    Altrimenti calcola dal DB.
    """
    override = _leggi_numero_config()
    if override:
        # Consuma l'override — non verrà più usato
        try:
            db_commit("UPDATE numero_fattura_config SET prossimo_numero=NULL WHERE id=1")
        except Exception:
            pass
        return override
    return _calc_prossimo_da_db()

def prossimo_numero_da_mostrare():
    """Per il label — mostra l'override se presente, altrimenti il calcolo dal DB."""
    override = _leggi_numero_config()
    return override if override else _calc_prossimo_da_db()

def imposta_numero_fattura_manuale():
    """Finestra per cambiare manualmente il prossimo numero fattura."""
    prossimo = prossimo_numero_da_mostrare()

    win = tk.Toplevel(root)
    win.title("Gestione Numero Fattura")
    win.geometry("420x210")
    win.grab_set(); win.resizable(False, False)

    tk.Label(win,
        text="Imposta il numero per la PROSSIMA fattura:",
        font=("Arial",11,"bold")).pack(pady=(16,2), padx=16, anchor="w")
    tk.Label(win,
        text=f"Numero automatico calcolato: {_calc_prossimo_da_db()}",
        fg="gray", font=("Arial",9)).pack(padx=16, anchor="w")

    frm = tk.Frame(win); frm.pack(padx=16, pady=12, fill="x")
    tk.Label(frm, text="Nuovo numero:", font=("Arial",10,"bold")).pack(side="left")
    e_num = tk.Entry(frm, width=12, font=("Arial",13,"bold"))
    e_num.insert(0, str(prossimo))
    e_num.select_range(0, "end")
    e_num.pack(side="left", padx=8); e_num.focus()

    tk.Label(win,
        text="Formato: AAMMNN  (es. 260401 = anno 26, aprile, n°01)\n"
             "Il numero inserito sarà usato per la prossima fattura.",
        fg="#666", font=("Arial",9), justify="left").pack(padx=16, anchor="w")

    def salva():
        try:
            nuovo = int(e_num.get().strip())
            if nuovo < 100101:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Errore",
                "Numero non valido.\nFormato: AAMMNN (es. 260401)", parent=win)
            return
        try:
            db_commit(
                "INSERT OR REPLACE INTO numero_fattura_config (id, prossimo_numero) VALUES (1, ?)",
                (nuovo,))
        except Exception:
            db_commit(
                "CREATE TABLE IF NOT EXISTS numero_fattura_config "
                "(id INT PRIMARY KEY DEFAULT 1, prossimo_numero INT NULL)")
            db_commit(
                "INSERT OR REPLACE INTO numero_fattura_config (id, prossimo_numero) VALUES (1, ?)",
                (nuovo,))
        messagebox.showinfo("✅ Impostato",
            f"La prossima fattura sarà: {nuovo}", parent=win)
        win.destroy()
        aggiorna_lbl_num_fat()

    tk.Button(win, text="💾 Imposta numero",
              command=salva, bg="#1565C0", fg="white",
              font=("Arial",10,"bold"), width=22).pack(pady=8)
    win.bind("<Return>", lambda e: salva())


# ── STAMPA testo su stampante di sistema ──────────────────────
def stampa_testo(titolo, testo):
    """Apre una finestra di anteprima con pulsante Stampa."""
    win = tk.Toplevel()
    win.title(f"Stampa – {titolo}"); win.geometry("700x600"); win.grab_set()
    tk.Label(win, text=titolo, font=("Arial",12,"bold")).pack(pady=6)
    frame_txt = tk.Frame(win); frame_txt.pack(fill="both", expand=True, padx=10, pady=4)
    t = tk.Text(frame_txt, font=("Courier",9), wrap="none")
    sb_v = ttk.Scrollbar(frame_txt, orient="vertical",   command=t.yview)
    sb_h = ttk.Scrollbar(frame_txt, orient="horizontal", command=t.xview)
    t.config(yscrollcommand=sb_v.set, xscrollcommand=sb_h.set)
    t.grid(row=0,column=0,sticky="nsew"); sb_v.grid(row=0,column=1,sticky="ns"); sb_h.grid(row=1,column=0,sticky="ew")
    frame_txt.rowconfigure(0,weight=1); frame_txt.columnconfigure(0,weight=1)
    t.insert("1.0", testo); t.config(state="disabled")

    def stampa():
        import tempfile
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".txt", mode="w", encoding="utf-8")
        tmp.write(testo); tmp.close()
        if sys.platform == "win32":
            os.startfile(tmp.name, "print")
        else:
            subprocess.call(["lpr", tmp.name])
        messagebox.showinfo("Stampa","Documento inviato alla stampante.", parent=win)

    tk.Button(win, text="🖨 Stampa", command=stampa, width=16,
              bg="#1565C0", fg="white", font=("Arial",10,"bold")).pack(pady=8)

# ── SCANNER via WIA (Windows) ─────────────────────────────────
def scansiona_immagine(parent_win, callback_path):
    """
    Acquisisce da scanner usando WIA nativo di Windows.
    Mostra il dialogo di selezione dispositivo nativo di Windows.
    Chiama callback_path(percorso_file) se riesce.
    """
    import tempfile
    out_path = os.path.join(tempfile.gettempdir(),
                            f"scan_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.jpg")
    out_escaped = out_path.replace("\\", "\\\\")

    # DeviceType 0 = UnspecifiedDeviceType  (mostra TUTTI i dispositivi WIA, inclusi scanner)
    # Intent 1 = Color,  Bias 64 = MaxQuality
    # FormatGUID JPEG = {B96B3CAE-0728-11D3-9D7B-0000F81EF32E}
    ps = f"""
$out = "{out_escaped}"
try {{
    $dlg = New-Object -ComObject WIA.CommonDialog
    $img = $dlg.ShowAcquireImage(0, 1, 64, "{{B96B3CAE-0728-11D3-9D7B-0000F81EF32E}}", $false, $true, $false)
    if ($img -eq $null) {{
        Write-Host "ANNULLATO"
        exit 0
    }}
    $img.SaveFile($out)
    Write-Host "OK:$out"
}} catch {{
    Write-Host "ERR:$($_.Exception.Message)"
    exit 1
}}
"""
    try:
        r = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps],
            capture_output=True, text=True, timeout=180)

        output = (r.stdout or "").strip()
        last   = output.splitlines()[-1] if output else ""

        if last == "ANNULLATO":
            return  # utente ha premuto Annulla — nessun errore

        if last.startswith("ERR:") or r.returncode != 0:
            msg = last.replace("ERR:", "") if last.startswith("ERR:") else \
                  (r.stderr or "Errore sconosciuto").strip().splitlines()[-1]
            raise Exception(msg)

        if not os.path.exists(out_path):
            raise Exception("File scansionato non trovato. Riprova.")

        callback_path(out_path)

    except subprocess.TimeoutExpired:
        messagebox.showwarning("Scanner", "Operazione annullata (timeout).", parent=parent_win)
    except Exception as ex:
        err_msg = str(ex)
        # Messaggio più chiaro per errore comune "no device"
        if "tipo selezionato" in err_msg or "WiaDeviceType" in err_msg or "no device" in err_msg.lower():
            err_msg = ("Nessuno scanner trovato da Windows.\n\n"
                       "Cosa fare:\n"
                       "1. Collega lo scanner via USB e accendilo\n"
                       "2. Installa il driver dallo scanner (CD o sito produttore)\n"
                       "3. Verifica in Pannello di Controllo → Scanner e Fotocamere\n\n"
                       "Puoi comunque usare 'Sfoglia file' per allegare\n"
                       "un'immagine già salvata sul PC.")
        messagebox.showerror("Scanner non disponibile", err_msg, parent=parent_win)

# ══════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════
# SISTEMA DI LOGIN
# ══════════════════════════════════════════════════════════════
import hashlib

CREDENZIALI_FILE = os.path.join(BASE_DIR, "credenziali.dat")

def _hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def _leggi_credenziali():
    try:
        if os.path.exists(CREDENZIALI_FILE):
            with open(CREDENZIALI_FILE, "r") as f:
                lines = f.read().strip().splitlines()
                if len(lines) >= 2:
                    return lines[0], lines[1]  # username, hash_password
    except Exception:
        pass
    return None, None

def _salva_credenziali(username, hash_pwd):
    with open(CREDENZIALI_FILE, "w") as f:
        f.write(f"{username}\n{hash_pwd}")

def _finestra_primo_avvio():
    """Primo avvio: chiede di creare username e password."""
    win = tk.Toplevel()
    win.title("Benvenuto – Imposta accesso")
    win.geometry("400x280")
    win.resizable(False, False)
    win.grab_set()
    win.protocol("WM_DELETE_WINDOW", lambda: None)  # non chiudibile

    tk.Label(win, text="🔧 Garage Tito – Primo accesso",
             font=("Arial",13,"bold")).pack(pady=(20,4))
    tk.Label(win, text="Crea le tue credenziali di accesso:",
             font=("Arial",10), fg="#555").pack(pady=(0,12))

    frm = tk.Frame(win); frm.pack()
    tk.Label(frm, text="Nome utente:", width=14, anchor="e").grid(row=0, column=0, padx=8, pady=6)
    e_user = tk.Entry(frm, width=22, font=("Arial",11))
    e_user.grid(row=0, column=1, padx=8, pady=6)

    tk.Label(frm, text="Password:", width=14, anchor="e").grid(row=1, column=0, padx=8, pady=6)
    e_pwd = tk.Entry(frm, width=22, font=("Arial",11), show="●")
    e_pwd.grid(row=1, column=1, padx=8, pady=6)

    tk.Label(frm, text="Conferma password:", width=14, anchor="e").grid(row=2, column=0, padx=8, pady=6)
    e_pwd2 = tk.Entry(frm, width=22, font=("Arial",11), show="●")
    e_pwd2.grid(row=2, column=1, padx=8, pady=6)

    lbl_err = tk.Label(win, text="", fg="red", font=("Arial",9))
    lbl_err.pack()

    def conferma():
        u = e_user.get().strip()
        p = e_pwd.get()
        p2 = e_pwd2.get()
        if not u:
            lbl_err.config(text="Inserisci un nome utente."); return
        if len(p) < 4:
            lbl_err.config(text="Password troppo corta (min 4 caratteri)."); return
        if p != p2:
            lbl_err.config(text="Le password non coincidono."); return
        _salva_credenziali(u, _hash_password(p))
        win.destroy()

    tk.Button(win, text="✅ Crea accesso", command=conferma,
              bg="#1565C0", fg="white", font=("Arial",11,"bold"),
              width=18).pack(pady=10)
    e_user.focus()
    win.bind("<Return>", lambda e: conferma())
    win.wait_window()

def _finestra_login():
    """Finestra di login standard."""
    win = tk.Toplevel()
    win.title("Accesso – Garage Tito")
    win.geometry("360x220")
    win.resizable(False, False)
    win.grab_set()
    win.protocol("WM_DELETE_WINDOW", lambda: _chiudi_app())

    username_salvato, _ = _leggi_credenziali()

    tk.Label(win, text="🔑 Gestionale Garage Tito",
             font=("Arial",13,"bold")).pack(pady=(20,4))

    frm = tk.Frame(win); frm.pack(pady=8)
    tk.Label(frm, text="Utente:", width=12, anchor="e").grid(row=0, column=0, padx=8, pady=6)
    e_user = tk.Entry(frm, width=20, font=("Arial",11))
    e_user.insert(0, username_salvato or "")
    e_user.grid(row=0, column=1, padx=8, pady=6)

    tk.Label(frm, text="Password:", width=12, anchor="e").grid(row=1, column=0, padx=8, pady=6)
    e_pwd = tk.Entry(frm, width=20, font=("Arial",11), show="●")
    e_pwd.grid(row=1, column=1, padx=8, pady=6)

    lbl_err = tk.Label(win, text="", fg="red", font=("Arial",9))
    lbl_err.pack()

    accesso_ok = [False]

    def tenta_login():
        u = e_user.get().strip()
        p = e_pwd.get()
        u_sal, h_sal = _leggi_credenziali()
        if u == u_sal and _hash_password(p) == h_sal:
            accesso_ok[0] = True
            win.destroy()
        else:
            lbl_err.config(text="Utente o password errati.")
            e_pwd.delete(0, "end")
            e_pwd.focus()

    def _chiudi_app():
        win.destroy()
        import sys; sys.exit(0)

    tk.Button(win, text="🔓 Accedi", command=tenta_login,
              bg="#1565C0", fg="white", font=("Arial",11,"bold"),
              width=14).pack(pady=6)

    # Focus sul campo password se utente già compilato
    if username_salvato:
        e_pwd.focus()
    else:
        e_user.focus()
    win.bind("<Return>", lambda e: tenta_login())
    win.wait_window()
    return accesso_ok[0]

# ── Esegui login prima di aprire il gestionale ─────────────────
_login_root = tk.Tk()
_login_root.withdraw()  # nascondi finestra principale durante login

u_sal, h_sal = _leggi_credenziali()
if not u_sal:
    # Primo avvio: crea credenziali
    _finestra_primo_avvio()
else:
    # Login normale
    if not _finestra_login():
        _login_root.destroy()
        import sys; sys.exit(0)

_login_root.destroy()

root = tk.Tk()
root.title("Gestionale Officina – Garage Tito")
root.resizable(True, True)
# Adatta la finestra allo schermo disponibile
_sw = root.winfo_screenwidth()
_sh = root.winfo_screenheight()
_w  = min(1200, int(_sw * 0.92))
_h  = min(800,  int(_sh * 0.90))
root.geometry(f"{_w}x{_h}+{(_sw-_w)//2}+{(_sh-_h)//2}")

style = ttk.Style()
style.configure("Treeview", font=("Arial",10), rowheight=26)
style_sto = ttk.Style()
style_sto.configure("Storico.Treeview", font=("Arial",10), rowheight=52)
style.configure("Treeview.Heading", font=("Arial",10,"bold"))

# Schede colorate e più grandi — usa tema "clam" che supporta i colori
try:
    style.theme_use("clam")
except Exception:
    pass

style.configure("TNotebook", background="#d0d0d0", tabmargins=[2,2,2,0])
style.configure("TNotebook.Tab",
    font=("Arial", 11, "bold"),
    padding=[16, 9],
    background="#d0d0d0",
    foreground="#333333")
style.map("TNotebook.Tab",
    background=[("selected","#1565C0"), ("active","#4a90d9"), ("!selected","#d0d0d0")],
    foreground=[("selected","white"),   ("active","white"),   ("!selected","#555555")],
    expand=[("selected",[1,1,1,0])])

nb = ttk.Notebook(root)
nb.pack(fill="both", expand=True, padx=8, pady=8)

# ══════════════════════════════════════════════════════════════
# SCHEDA 1 – CLIENTI & VEICOLI
# ══════════════════════════════════════════════════════════════
tab_clienti = tk.Frame(nb)
nb.add(tab_clienti, text="👤  Clienti & Veicoli")

frame_lista = tk.LabelFrame(tab_clienti, text="Clienti", width=270)
frame_lista.pack(side="left", fill="y", padx=8, pady=8)
frame_lista.pack_propagate(False)

search_var = tk.StringVar()
tk.Label(frame_lista, text="🔍 Cerca (cognome o targa):").pack(anchor="w", padx=4, pady=(4,0))
tk.Entry(frame_lista, textvariable=search_var).pack(fill="x", padx=4, pady=(0,4))

lista_clienti = tk.Listbox(frame_lista, width=30, font=("Arial",10))
sc_lc = ttk.Scrollbar(frame_lista, orient="vertical", command=lista_clienti.yview)
lista_clienti.config(yscrollcommand=sc_lc.set)
lista_clienti.pack(side="left", fill="both", expand=True, padx=(4,0), pady=4)
sc_lc.pack(side="left", fill="y", pady=4)

clienti_ids = []

def aggiorna_lista_clienti(*_):
    global clienti_ids
    filtro = search_var.get().strip()
    if filtro:
        f = f"%{filtro}%"
        rows = db_fetch(
            "SELECT DISTINCT c.id_cliente, c.cognome, c.nome, c.tipo_cliente, c.ragione_sociale "
            "FROM clienti c "
            "LEFT JOIN veicoli v ON v.id_cliente = c.id_cliente "
            "WHERE CONCAT(c.cognome,' ',c.nome) LIKE %s "
            "   OR IFNULL(c.ragione_sociale,'') LIKE %s "
            "   OR v.targa LIKE %s "
            "ORDER BY c.cognome, c.nome", (f, f, f))
    else:
        rows = db_fetch(
            "SELECT id_cliente, cognome, nome, tipo_cliente, ragione_sociale "
            "FROM clienti ORDER BY cognome, nome")
    lista_clienti.delete(0,"end"); clienti_ids = []
    for r in rows:
        if r.get("tipo_cliente") == "azienda":
            label = f"🏢 {r['ragione_sociale'] or r['cognome']}"
        else:
            label = f"{r['cognome']} {r['nome']}"
        lista_clienti.insert("end", label)
        clienti_ids.append(r["id_cliente"])

search_var.trace_add("write", aggiorna_lista_clienti)

# pannello destro
frame_det = tk.Frame(tab_clienti)
frame_det.pack(side="left", fill="both", expand=True, padx=8, pady=8)

lf_dati = tk.LabelFrame(frame_det, text="Dati Cliente")
lf_dati.pack(fill="x")

# Layout: riga 0=Nome+Cognome, riga 1=Indirizzo+CAP, riga 2=Telefono+Email
det_labels = ["Nome","Cognome","Indirizzo","CAP","Telefono","Email"]
det_vars   = {l: tk.StringVar() for l in det_labels}
disposizione = [("Nome","Cognome"),("Indirizzo","CAP"),("Telefono","Email")]
for riga, (lbl_sx, lbl_dx) in enumerate(disposizione):
    tk.Label(lf_dati, text=lbl_sx, width=9, anchor="e").grid(row=riga, column=0, padx=6, pady=4, sticky="e")
    tk.Entry(lf_dati, textvariable=det_vars[lbl_sx], width=24, state="readonly").grid(row=riga, column=1, padx=4, pady=4, sticky="w")
    tk.Label(lf_dati, text=lbl_dx, width=7, anchor="e").grid(row=riga, column=2, padx=6, pady=4, sticky="e")
    tk.Entry(lf_dati, textvariable=det_vars[lbl_dx], width=20, state="readonly").grid(row=riga, column=3, padx=4, pady=4, sticky="w")

lf_veicoli = tk.LabelFrame(frame_det, text="Veicoli del Cliente  (doppio clic = modifica)")
lf_veicoli.pack(fill="both", expand=True, pady=6)

cols_v = ("targa","marca","modello","anno","km")
tree_v = ttk.Treeview(lf_veicoli, columns=cols_v, show="headings", height=7)
for c,w,h in zip(cols_v,(90,100,140,65,80),("Targa","Marca","Modello","Anno","Km")):
    tree_v.heading(c,text=h); tree_v.column(c,width=w)
sc_v = ttk.Scrollbar(lf_veicoli, orient="vertical", command=tree_v.yview)
tree_v.config(yscrollcommand=sc_v.set)
tree_v.pack(side="left", fill="both", expand=True, padx=4, pady=4)
sc_v.pack(side="right", fill="y")

current_cliente_id = None

def on_select_cliente(event=None):
    global current_cliente_id
    sel = lista_clienti.curselection()
    if not sel: return
    cid = clienti_ids[sel[0]]; current_cliente_id = cid
    rows = db_fetch("SELECT * FROM clienti WHERE id_cliente=%s",(cid,))
    if not rows: return
    r = rows[0]
    for lbl,key in [("Nome","nome"),("Cognome","cognome"),("Telefono","telefono"),
                    ("Email","email"),("Indirizzo","indirizzo"),("CAP","cap")]:
        det_vars[lbl].set(r[key] or "")
    # Se è azienda, mostra ragione sociale e tipo
    if r.get("tipo_cliente") == "azienda":
        det_vars["Nome"].set(r.get("ragione_sociale") or r.get("cognome") or "")
        det_vars["Cognome"].set(f"🏢 Azienda  |  P.IVA/CF: {r.get('nome') or '—'}")
    carica_veicoli_tree(cid)

def carica_veicoli_tree(cid):
    tree_v.delete(*tree_v.get_children())
    for r in db_fetch("SELECT id_veicolo,targa,marca,modello,anno,chilometri FROM veicoli WHERE id_cliente=%s",(cid,)):
        tree_v.insert("","end", iid=str(r["id_veicolo"]),
            values=(r["targa"] or "",r["marca"] or "",r["modello"] or "",r["anno"] or "",r["chilometri"] or 0))

lista_clienti.bind("<<ListboxSelect>>", on_select_cliente)

# ── finestre cliente ──────────────────────────────────────────
def _finestra_cliente(titolo, dati_iniziali=None, on_salva=None):
    """Finestra generica per aggiungere o modificare un cliente/azienda."""
    win = tk.Toplevel(root)
    win.title(titolo); win.geometry("370x400"); win.grab_set(); win.resizable(False, False)

    is_azienda = tk.BooleanVar(value=(dati_iniziali or {}).get("tipo_cliente") == "azienda")

    # Riga tipo cliente
    top = tk.Frame(win); top.pack(fill="x", padx=12, pady=(12,4))
    tk.Label(top, text="Tipo:", font=("Arial",10,"bold")).pack(side="left")
    tk.Radiobutton(top, text="👤 Privato", variable=is_azienda, value=False,
                   command=lambda: _aggiorna_layout()).pack(side="left", padx=8)
    tk.Radiobutton(top, text="🏢 Azienda", variable=is_azienda, value=True,
                   command=lambda: _aggiorna_layout()).pack(side="left", padx=8)

    frm = tk.Frame(win); frm.pack(fill="both", expand=True, padx=12)

    entries = {}

    def _campo(row, lbl, key, valore=""):
        tk.Label(frm, text=lbl, anchor="e", width=16).grid(row=row, column=0, padx=6, pady=5, sticky="e")
        e = tk.Entry(frm, width=26)
        e.insert(0, valore)
        e.grid(row=row, column=1, padx=6, pady=5, sticky="w")
        entries[key] = e
        return e

    d = dati_iniziali or {}

    # Campi comuni a privato e azienda
    _campo(2, "Telefono", "telefono", d.get("telefono") or "")
    _campo(3, "Email",    "email",    d.get("email") or "")
    _campo(4, "Indirizzo","indirizzo",d.get("indirizzo") or "")
    _campo(5, "CAP",      "cap",      d.get("cap") or "")

    # Labels dinamiche per riga 0 e 1
    lbl0 = tk.Label(frm, text="", anchor="e", width=16)
    lbl0.grid(row=0, column=0, padx=6, pady=5, sticky="e")
    e0 = tk.Entry(frm, width=26)
    e0.grid(row=0, column=1, padx=6, pady=5, sticky="w")
    entries["_r0"] = e0

    lbl1 = tk.Label(frm, text="", anchor="e", width=16)
    lbl1.grid(row=1, column=0, padx=6, pady=5, sticky="e")
    e1 = tk.Entry(frm, width=26)
    e1.grid(row=1, column=1, padx=6, pady=5, sticky="w")
    entries["_r1"] = e1

    def _aggiorna_layout():
        if is_azienda.get():
            lbl0.config(text="Ragione Sociale*")
            lbl1.config(text="P.IVA / CF")
            entries["_r0"].delete(0,"end")
            entries["_r1"].delete(0,"end")
            entries["_r0"].insert(0, d.get("ragione_sociale") or "")
            entries["_r1"].insert(0, d.get("nome") or "")  # usiamo nome per P.IVA
            e1.config(state="normal")
        else:
            lbl0.config(text="Nome*")
            lbl1.config(text="Cognome*")
            entries["_r0"].delete(0,"end")
            entries["_r1"].delete(0,"end")
            entries["_r0"].insert(0, d.get("nome") or "")
            entries["_r1"].insert(0, d.get("cognome") or "")

    _aggiorna_layout()

    def salva():
        telefono  = entries["telefono"].get().strip() or None
        email     = entries["email"].get().strip() or None
        indirizzo = entries["indirizzo"].get().strip() or None
        cap       = entries["cap"].get().strip() or None

        if is_azienda.get():
            ragione = entries["_r0"].get().strip()
            piva    = entries["_r1"].get().strip() or None
            if not ragione:
                messagebox.showwarning("Errore","Ragione Sociale obbligatoria",parent=win); return
            # Per le aziende: cognome = ragione sociale (per compatibilità lista), nome = P.IVA
            on_salva(
                tipo="azienda",
                nome=piva or "",
                cognome=ragione,
                ragione_sociale=ragione,
                telefono=telefono, email=email,
                indirizzo=indirizzo, cap=cap)
        else:
            nome    = entries["_r0"].get().strip()
            cognome = entries["_r1"].get().strip()
            if not nome or not cognome:
                messagebox.showwarning("Errore","Nome e Cognome obbligatori",parent=win); return
            on_salva(
                tipo="privato",
                nome=nome, cognome=cognome,
                ragione_sociale=None,
                telefono=telefono, email=email,
                indirizzo=indirizzo, cap=cap)
        win.destroy()

    tk.Button(win, text="💾 Salva", command=salva, width=20,
              bg="#4CAF50", fg="white", font=("Arial",10,"bold")).pack(pady=12)
    win.bind("<Return>", lambda e: salva())

def finestra_aggiungi_cliente():
    def on_salva(tipo, nome, cognome, ragione_sociale, telefono, email, indirizzo, cap):
        db_commit(
            "INSERT INTO clienti (tipo_cliente,nome,cognome,ragione_sociale,telefono,email,indirizzo,cap) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
            (tipo, nome, cognome, ragione_sociale, telefono, email, indirizzo, cap))
        messagebox.showinfo("✅ OK", "Cliente aggiunto!")
        aggiorna_lista_clienti()
    _finestra_cliente("Nuovo Cliente", on_salva=on_salva)

def finestra_modifica_cliente():
    if not current_cliente_id:
        messagebox.showwarning("Attenzione","Seleziona prima un cliente"); return
    r = db_fetch("SELECT * FROM clienti WHERE id_cliente=%s",(current_cliente_id,))[0]
    def on_salva(tipo, nome, cognome, ragione_sociale, telefono, email, indirizzo, cap):
        db_commit(
            "UPDATE clienti SET tipo_cliente=%s,nome=%s,cognome=%s,ragione_sociale=%s,"
            "telefono=%s,email=%s,indirizzo=%s,cap=%s WHERE id_cliente=%s",
            (tipo, nome, cognome, ragione_sociale, telefono, email, indirizzo, cap, current_cliente_id))
        messagebox.showinfo("✅ OK","Cliente aggiornato!")
        aggiorna_lista_clienti(); on_select_cliente()
    _finestra_cliente("Modifica Cliente", dati_iniziali=r, on_salva=on_salva)

def elimina_cliente():
    if not current_cliente_id:
        messagebox.showwarning("Attenzione","Seleziona prima un cliente"); return
    r = db_fetch("SELECT cognome,nome,tipo_cliente,ragione_sociale FROM clienti WHERE id_cliente=%s",(current_cliente_id,))[0]
    if r.get("tipo_cliente") == "azienda":
        nome_completo = f"🏢 {r.get('ragione_sociale') or r['cognome']}"
    else:
        nome_completo = f"{r['cognome']} {r['nome']}"
    veicoli = db_fetch("SELECT COUNT(*) AS n FROM veicoli WHERE id_cliente=%s",(current_cliente_id,))[0]["n"]
    msg = (f"Eliminare il cliente  {nome_completo}?\n\n"
           f"Verranno eliminati anche:\n"
           f"  • {veicoli} veicolo/i con relative carte grige\n"
           f"  • Tutto lo storico interventi\n"
           f"  • Le fatture associate\n\n"
           f"Questa operazione NON può essere annullata.")
    if not messagebox.askyesno("⚠️ Conferma eliminazione", msg, icon="warning"): return
    db_commit("DELETE FROM clienti WHERE id_cliente=%s",(current_cliente_id,))
    messagebox.showinfo("OK",f"Cliente {nome_completo} eliminato.")
    aggiorna_lista_clienti()
    for lbl in det_labels: det_vars[lbl].set("")
    tree_v.delete(*tree_v.get_children())

def stampa_clienti():
    rows = db_fetch("SELECT cognome,nome,indirizzo,cap,email,telefono FROM clienti ORDER BY cognome,nome")
    if not rows:
        messagebox.showinfo("Stampa","Nessun cliente da stampare."); return
    # Calcola larghezze dinamiche basate sul contenuto reale + minimo garantito
    W_NOM = max(20, max(len(r['cognome']+' '+r['nome']) for r in rows) + 3)
    W_IND = max(22, max(len((r['indirizzo'] or '')+" "+(r['cap'] or '')) for r in rows) + 3)
    W_EML = max(22, max(len(r['email'] or '') for r in rows) + 3)
    totale = W_NOM + W_IND + W_EML + 16
    sep = "─" * totale
    intestazione = f"{'Cognome Nome':<{W_NOM}} {'Indirizzo + CAP':<{W_IND}} {'Email':<{W_EML}} Telefono"
    lines = [f"{'ELENCO CLIENTI':^{totale}}", f"{'Garage Tito':^{totale}}", sep, intestazione, sep]
    for r in rows:
        ind = f"{r['indirizzo'] or ''} {r['cap'] or ''}".strip()
        nome = r['cognome']+' '+r['nome']
        lines.append(f"{nome:<{W_NOM}} {ind:<{W_IND}} {r['email'] or '':<{W_EML}} {r['telefono'] or ''}")
    lines += [sep, f"Totale clienti: {len(rows)}   –   Stampato il {datetime.date.today().strftime('%d/%m/%Y')}"]
    stampa_testo("Elenco Clienti", "\n".join(lines))

# ── finestre veicolo ──────────────────────────────────────────
def _finestra_veicolo(vid_esistente=None):
    """Usata sia per Aggiungi che per Modifica veicolo."""
    if not current_cliente_id:
        messagebox.showwarning("Attenzione","Seleziona prima un cliente"); return
    dati = None
    if vid_esistente:
        rows = db_fetch("SELECT * FROM veicoli WHERE id_veicolo=%s",(vid_esistente,))
        if not rows: return
        dati = rows[0]
    win = tk.Toplevel(root)
    win.title("Modifica Veicolo" if dati else "Aggiungi Veicolo")
    win.geometry("430x520"); win.grab_set()
    campi = [("Targa*","targa"),("Marca","marca"),("Modello","modello"),
             ("Anno","anno"),("Chilometri","chilometri"),("Telaio","telaio")]
    entries = {}
    for i,(lbl,key) in enumerate(campi):
        tk.Label(win,text=lbl).grid(row=i,column=0,padx=10,pady=5,sticky="e")
        e = tk.Entry(win,width=24); e.grid(row=i,column=1,padx=10,pady=5,sticky="w")
        if dati: e.insert(0, str(dati[key] or ""))
        elif key=="chilometri": e.insert(0,"0")
        entries[key]=e
    entries["targa"].focus()

    lf_carta = tk.LabelFrame(win,text="Carta Grigia",padx=8,pady=6)
    lf_carta.grid(row=len(campi),column=0,columnspan=2,padx=10,pady=6,sticky="ew")
    carta_path = [dati["carta_grigia_path"] if dati else None]
    testo_carta = os.path.basename(dati["carta_grigia_path"]) if dati and dati["carta_grigia_path"] else "Nessuna carta grigia"
    colore_carta = "green" if (dati and dati["carta_grigia_path"]) else "gray"
    lbl_carta = tk.Label(lf_carta, text=f"{'✅ ' if colore_carta=='green' else ''}{testo_carta}",
                         fg=colore_carta, wraplength=340)
    lbl_carta.pack(fill="x",pady=(0,6))
    btn_row = tk.Frame(lf_carta); btn_row.pack(fill="x")

    def scegli():
        fp = filedialog.askopenfilename(parent=win,title="Seleziona carta grigia",
            filetypes=[("Immagini/PDF","*.jpg *.jpeg *.png *.pdf"),("Tutti","*.*")])
        if fp:
            carta_path[0] = fp
            lbl_carta.config(text=f"✅ {os.path.basename(fp)}", fg="green")

    def scan_cb(fp):
        carta_path[0] = fp
        lbl_carta.config(text=f"✅ Scansione: {os.path.basename(fp)}", fg="green")

    def scansiona():
        lbl_carta.config(text="⏳ Acquisizione in corso…", fg="orange")
        win.update()
        scansiona_immagine(win, scan_cb)

    tk.Button(btn_row,text="📎 Sfoglia file…",command=scegli,   width=16).pack(side="left",padx=4)
    tk.Button(btn_row,text="🖨 Scansiona ora…",command=scansiona,width=16).pack(side="left",padx=4)

    def salva():
        targa = entries["targa"].get().strip().upper()
        if not targa:
            messagebox.showwarning("Errore","La targa è obbligatoria",parent=win); return
        try:
            anno = int(entries["anno"].get()) if entries["anno"].get().strip() else None
            km   = int(entries["chilometri"].get()) if entries["chilometri"].get().strip() else 0
        except ValueError:
            messagebox.showwarning("Errore","Anno e Chilometri devono essere numeri interi",parent=win); return
        # gestisci carta grigia
        rel = dati["carta_grigia_path"] if dati else None  # mantieni quella esistente se non cambiata
        if carta_path[0] and carta_path[0] != (dati["carta_grigia_path"] if dati else None):
            nome_f = os.path.basename(carta_path[0])
            shutil.copy(carta_path[0], os.path.join(DIR_DOCS, nome_f))
            rel = f"documenti/{nome_f}"
        if dati:
            db_commit("UPDATE veicoli SET targa=%s,marca=%s,modello=%s,anno=%s,chilometri=%s,telaio=%s,carta_grigia_path=%s WHERE id_veicolo=%s",
                (targa,entries["marca"].get().strip() or None,entries["modello"].get().strip() or None,
                 anno,km,entries["telaio"].get().strip() or None,rel,vid_esistente))
            messagebox.showinfo("OK","Veicolo aggiornato!",parent=win)
        else:
            db_commit("INSERT INTO veicoli (id_cliente,targa,marca,modello,anno,chilometri,telaio,carta_grigia_path) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                (current_cliente_id,targa,entries["marca"].get().strip() or None,
                 entries["modello"].get().strip() or None,anno,km,
                 entries["telaio"].get().strip() or None,rel))
            messagebox.showinfo("OK","Veicolo aggiunto!",parent=win)
        win.destroy(); carica_veicoli_tree(current_cliente_id)

    tk.Button(win,text="✅  Salva Veicolo",command=salva,width=18,
              bg="#4CAF50",fg="white",font=("Arial",10,"bold")).grid(row=len(campi)+1,column=0,columnspan=2,pady=10)
    win.columnconfigure(1,weight=1)

def finestra_aggiungi_veicolo(): _finestra_veicolo()

def finestra_modifica_veicolo():
    sel = tree_v.selection()
    if not sel:
        messagebox.showwarning("Attenzione","Seleziona un veicolo dalla lista"); return
    _finestra_veicolo(int(sel[0]))

def on_doppio_clic_veicolo(event):
    _finestra_veicolo(int(tree_v.selection()[0])) if tree_v.selection() else None

tree_v.bind("<Double-1>", on_doppio_clic_veicolo)

def elimina_veicolo():
    sel = tree_v.selection()
    if not sel:
        messagebox.showwarning("Attenzione","Seleziona un veicolo dalla lista"); return
    vid = int(sel[0])
    r = db_fetch("SELECT targa,marca,modello FROM veicoli WHERE id_veicolo=%s",(vid,))[0]
    desc = f"{r['targa']} – {r['marca'] or ''} {r['modello'] or ''}".strip()
    if not messagebox.askyesno("Conferma",f"Eliminare il veicolo {desc}?\n\nVerrà eliminato anche tutto lo storico interventi.", icon="warning"): return
    db_commit("DELETE FROM veicoli WHERE id_veicolo=%s",(vid,))
    carica_veicoli_tree(current_cliente_id)

def vedi_carta_grigia():
    sel = tree_v.selection()
    if not sel:
        messagebox.showwarning("Attenzione","Seleziona un veicolo"); return
    rows = db_fetch("SELECT carta_grigia_path FROM veicoli WHERE id_veicolo=%s",(int(sel[0]),))
    path = rows[0]["carta_grigia_path"] if rows else None
    if not path:
        messagebox.showinfo("Carta grigia","Nessuna carta grigia salvata per questo veicolo"); return
    full = os.path.join(BASE_DIR, path)
    if not os.path.exists(full):
        messagebox.showwarning("Attenzione",f"File non trovato:\n{full}"); return
    apri_file(full)

# pulsanti scheda clienti
btn_frame = tk.Frame(frame_det); btn_frame.pack(fill="x", pady=9)
tk.Button(btn_frame,text="➕ Nuovo Cliente",    width=15, command=finestra_aggiungi_cliente, font=("Arial",11,"bold")).pack(side="left",padx=10)
tk.Button(btn_frame,text="✏️ Modifica Cliente",  width=15, command=finestra_modifica_cliente, font=("Arial",11,"bold")).pack(side="left",padx=10)
tk.Button(btn_frame,text="🗑 Elimina Cliente",   width=15, command=elimina_cliente,           font=("Arial",11,"bold")).pack(side="left",padx=10)
tk.Button(btn_frame,text="🖨 Stampa Clienti",    width=15, command=stampa_clienti,            font=("Arial",11,"bold")).pack(side="left",padx=10)
tk.Button(btn_frame,text="🚗 Aggiungi Veicolo",  width=15, command=finestra_aggiungi_veicolo, font=("Arial",11,"bold")).pack(side="left",padx=10)
tk.Button(btn_frame,text="✏️ Modifica Veicolo",  width=15, command=finestra_modifica_veicolo, font=("Arial",11,"bold")).pack(side="left",padx=10)
tk.Button(btn_frame,text="🗑 Elimina Veicolo",   width=15, command=elimina_veicolo,           font=("Arial",11,"bold")).pack(side="left",padx=10)
tk.Button(btn_frame,text="📄 Apri Carta Grigia", width=15, command=vedi_carta_grigia,         font=("Arial",11,"bold")).pack(side="left",padx=10)


# ══════════════════════════════════════════════════════════════
# SCHEDA 2 – FATTURE
# ══════════════════════════════════════════════════════════════
tab_fatture = tk.Frame(nb)
nb.add(tab_fatture, text="🧾  Fatture")

# Barra superiore con numero fattura corrente
num_bar = tk.Frame(tab_fatture, bg="#e8f0fe", relief="ridge", bd=1)
num_bar.pack(fill="x", padx=10, pady=(8,2))
lbl_num_fat = tk.Label(num_bar, text="Prossima fattura: —",
                       font=("Arial",10,"bold"), bg="#e8f0fe", fg="#1565C0")
lbl_num_fat.pack(side="left", padx=12, pady=4)

def aggiorna_lbl_num_fat():
    n = prossimo_numero_da_mostrare()
    lbl_num_fat.config(text=f"Prossima fattura N°: {n}")

def apri_gestione_numero():
    imposta_numero_fattura_manuale()
    aggiorna_lbl_num_fat()

tk.Button(num_bar, text="✏️ Cambia numero",
          command=apri_gestione_numero,
          font=("Arial",9,"bold"), bg="#1565C0", fg="white",
          relief="flat", padx=10).pack(side="right", padx=8, pady=4)

sel_frame = tk.LabelFrame(tab_fatture,text="Seleziona Cliente e Veicolo")
sel_frame.pack(fill="x",padx=10,pady=8)
tk.Label(sel_frame,text="Cliente:").grid(row=0,column=0,padx=6,pady=5,sticky="e")
fat_cli_var = tk.StringVar()
fat_cli_combo = ttk.Combobox(sel_frame,textvariable=fat_cli_var,width=30)
fat_cli_combo.grid(row=0,column=1,padx=6,pady=5)
tk.Label(sel_frame,text="Veicolo:").grid(row=0,column=2,padx=6,pady=5,sticky="e")
fat_vei_var = tk.StringVar()
fat_vei_combo = ttk.Combobox(sel_frame,textvariable=fat_vei_var,state="readonly",width=28)
fat_vei_combo.grid(row=0,column=3,padx=6,pady=5)
tk.Label(sel_frame,text="(scrivi per filtrare per cognome)",
         fg="gray",font=("Arial",8)).grid(row=1,column=0,columnspan=2,padx=6,sticky="w")

fat_cli_dict = {}; fat_vei_dict = {}
fat_cli_tutti = []   # lista completa per il filtro

def carica_fat_clienti():
    global fat_cli_tutti
    fat_cli_dict.clear(); fat_cli_tutti = []
    for r in db_fetch("SELECT id_cliente,cognome,nome FROM clienti ORDER BY cognome,nome"):
        k = f"{r['cognome']} {r['nome']}"
        fat_cli_dict[k] = r["id_cliente"]
        fat_cli_tutti.append(k)
    fat_cli_combo["values"] = fat_cli_tutti

def _fat_filtra_clienti(event=None):
    testo = fat_cli_var.get().lower()
    filtrati = [k for k in fat_cli_tutti if testo in k.lower()]
    fat_cli_combo["values"] = filtrati
    if fat_cli_var.get() in fat_cli_dict:
        on_fat_cli_change()

def on_fat_cli_change(event=None):
    cid = fat_cli_dict.get(fat_cli_var.get())
    fat_vei_dict.clear(); fat_vei_combo.set(""); fat_vei_combo["values"] = []
    if not cid: return
    for r in db_fetch("SELECT id_veicolo,targa,marca,modello FROM veicoli WHERE id_cliente=%s",(cid,)):
        fat_vei_dict[f"{r['targa']} – {r['marca'] or ''} {r['modello'] or ''}".strip(" –")] = r["id_veicolo"]
    fat_vei_combo["values"] = list(fat_vei_dict.keys())
    if len(fat_vei_dict) == 1:
        fat_vei_combo.current(0)

fat_cli_combo.bind("<KeyRelease>", _fat_filtra_clienti)
fat_cli_combo.bind("<<ComboboxSelected>>", on_fat_cli_change)

# Descrizione lavoro: va nella riga B17 del template Excel (prima delle voci)
lf_desc_lav = tk.LabelFrame(tab_fatture, text="Descrizione Lavoro  –  appare in cima alla fattura prima delle voci  (solo testo, niente prezzo)")
lf_desc_lav.pack(fill="x", padx=10, pady=(0,4))
desc_lav_txt = tk.Text(lf_desc_lav, height=2, font=("Arial",10), wrap="word")
desc_lav_txt.pack(fill="x", padx=6, pady=4)

lf_righe = tk.LabelFrame(tab_fatture,text="Voci Fattura  –  aggiungi le righe poi clicca GENERA FATTURA")
lf_righe.pack(fill="both",expand=True,padx=10,pady=4)
cols_r = ("descrizione","quantita","prezzo","totale_riga")
tree_r = ttk.Treeview(lf_righe,columns=cols_r,show="headings",height=12)
for c,w,h in zip(cols_r,(340,80,100,100),("Descrizione","Quantità","Prezzo CHF","Totale")):
    tree_r.heading(c,text=h); tree_r.column(c,width=w)
tree_r.pack(fill="both",expand=True,padx=4,pady=4)
righe_fattura = []

def aggiorna_tree_righe():
    tree_r.delete(*tree_r.get_children())
    for r in righe_fattura:
        tot = round(r["quantita"]*r["prezzo"],2)
        tree_r.insert("","end",values=(r["descrizione"],r["quantita"],f"{r['prezzo']:.2f}",f"{tot:.2f}"))
    lbl_totale.config(text=f"Totale imponibile: CHF {sum(r['quantita']*r['prezzo'] for r in righe_fattura):.2f}")

def aggiungi_riga():
    win = tk.Toplevel(root); win.title("Aggiungi voce fattura"); win.geometry("600x300"); win.grab_set(); win.resizable(True,True)
    tk.Label(win,text="Descrizione:",font=("Arial",10)).grid(row=0,column=0,padx=12,pady=10,sticky="ne")
    e_desc = tk.Text(win,width=52,height=4,font=("Arial",10),wrap="word")
    sb = ttk.Scrollbar(win,orient="vertical",command=e_desc.yview); e_desc.config(yscrollcommand=sb.set)
    e_desc.grid(row=0,column=1,padx=(12,0),pady=10,sticky="ew"); sb.grid(row=0,column=2,padx=(0,12),pady=10,sticky="ns"); e_desc.focus()
    tk.Label(win,text="Quantità:",font=("Arial",10)).grid(row=1,column=0,padx=12,pady=8,sticky="e")
    e_qty = tk.Entry(win,width=10,font=("Arial",10)); e_qty.insert(0,"1"); e_qty.grid(row=1,column=1,padx=12,pady=8,sticky="w")
    tk.Label(win,text="Prezzo CHF:",font=("Arial",10)).grid(row=2,column=0,padx=12,pady=8,sticky="e")
    e_prc = tk.Entry(win,width=10,font=("Arial",10)); e_prc.grid(row=2,column=1,padx=12,pady=8,sticky="w")
    def ok():
        desc = e_desc.get("1.0","end").strip()
        if not desc:
            messagebox.showwarning("Errore","Inserisci la descrizione",parent=win); return
        try: qty=float(e_qty.get().replace(",",".")); prc=float(e_prc.get().replace(",","."))
        except ValueError:
            messagebox.showwarning("Errore","Quantità e prezzo devono essere numeri",parent=win); return
        righe_fattura.append({"descrizione":desc,"quantita":qty,"prezzo":prc})
        aggiorna_tree_righe(); win.destroy()
    tk.Button(win,text="➕ Aggiungi voce",command=ok,width=18,bg="#4CAF50",fg="white",font=("Arial",10,"bold")).grid(row=3,column=0,columnspan=3,pady=12)
    win.columnconfigure(1,weight=1)

def rimuovi_riga():
    sel = tree_r.selection()
    if not sel: return
    righe_fattura.pop(tree_r.index(sel[0])); aggiorna_tree_righe()

def genera_fattura_excel():
    if not fat_cli_var.get() or not fat_vei_var.get():
        messagebox.showwarning("Errore","Seleziona cliente e veicolo"); return
    cid = fat_cli_dict.get(fat_cli_var.get())
    vid = fat_vei_dict.get(fat_vei_var.get())
    if not cid or not vid:
        messagebox.showwarning("Errore","Cliente o veicolo non valido"); return
    if not os.path.exists(TEMPLATE):
        messagebox.showerror("Errore",f"Template non trovato:\n{TEMPLATE}"); return

    cli = db_fetch("SELECT * FROM clienti WHERE id_cliente=%s",(cid,))[0]
    vei = db_fetch("SELECT * FROM veicoli WHERE id_veicolo=%s",(vid,))[0]
    desc_lav = desc_lav_txt.get("1.0","end").strip()
    totale_prev = round(sum(v["quantita"]*v["prezzo"] for v in righe_fattura),2)
    totale_prev_iva = round(totale_prev * (1 + IVA), 2)

    if cli.get("tipo_cliente") == "azienda":
        nome_cli = cli.get("ragione_sociale") or cli["cognome"]
    else:
        nome_cli = cli["cognome"] + " " + cli["nome"]
    targa_str = str(vei.get("targa",""))
    marca_str = (str(vei.get("marca","") or "") + " " + str(vei.get("modello","") or "")).strip()

    riepilogo = (
        "Cliente:  " + nome_cli + "\n"
        "Veicolo:  " + targa_str + " - " + marca_str + "\n"
        "Voci:     " + str(len(righe_fattura)) + " righe\n"
        "Totale:   CHF " + f"{totale_prev:.2f}" + " + IVA 8.1% = CHF " + f"{totale_prev_iva:.2f}" + "\n\n"
        "Confermi la creazione della fattura?")
    if not messagebox.askyesno("Conferma Fattura", riepilogo): return

    next_num = prossimo_numero_fattura()
    km_dlg = simpledialog.askinteger("Chilometri",
        "Km attuali del veicolo " + targa_str + "?\n(premi Annulla per saltare)",
        initialvalue=vei.get("chilometri") or 0, minvalue=0, maxvalue=9999999, parent=root)
    km_str = km_dlg if km_dlg is not None else (vei.get("chilometri") or "")
    if km_dlg is not None:
        db_commit("UPDATE veicoli SET chilometri=%s WHERE id_veicolo=%s",(km_dlg,vid))

    oggi_it   = datetime.date.today().strftime("%d.%m.%Y")
    oggi_file = datetime.date.today().strftime("%Y%m%d")
    nome_file = "Fattura_" + str(next_num) + "_" + cli["cognome"] + "_" + oggi_file + ".xlsm"
    dest      = os.path.join(DIR_FATT, nome_file)

    # ── Crea fattura con ZIP puro (preserva logo, VBA, stili) ──
    import zipfile as _zf, re as _re

    def _esc(s):
        return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

    try:
        with _zf.ZipFile(TEMPLATE, "r") as z:
            tmpl = {n: z.read(n) for n in z.namelist()
                    if n != "xl/calcChain.xml"}
        # Verifica che sia un xlsx valido
        if "[Content_Types].xml" not in tmpl:
            raise ValueError("Template non valido — manca [Content_Types].xml")

        s1 = tmpl["xl/worksheets/sheet1.xml"].decode("utf-8")
        ss = tmpl["xl/sharedStrings.xml"].decode("utf-8")
        s1_orig = s1  # copia originale per leggere gli stili

        str_list = _re.findall(r'<si><t(?:[^>]*)>(.*?)</t></si>', ss, _re.DOTALL)

        def _idx(val):
            e = _esc(val)
            if e in str_list: return str_list.index(e)
            str_list.append(e); return len(str_list)-1

        def _sty(ref):
            # Legge lo stile dall'originale (non modificato)
            m = _re.search(rf'<c r="{_re.escape(ref)}"([^>]*)>', s1_orig)
            if m:
                sm = _re.search(r's="(\d+)"', m.group(1))
                return f' s="{sm.group(1)}"' if sm else ""
            return ""

        def _pat(ref):
            r = _re.escape(ref)
            return rf'<c r="{r}"[^>]*/>|<c r="{r}"[^>]*>.*?</c>'

        def sstr(xml, ref, val):
            if not val: return xml
            new = f'<c r="{ref}"{_sty(ref)} t="s"><v>{_idx(val)}</v></c>'
            return _re.sub(_pat(ref), new, xml, flags=_re.DOTALL)

        def snum(xml, ref, val):
            new = f'<c r="{ref}"{_sty(ref)}><v>{val}</v></c>'
            return _re.sub(_pat(ref), new, xml, flags=_re.DOTALL)

        def sform(xml, ref, formula, val):
            new = f'<c r="{ref}"{_sty(ref)}><f>{formula}</f><v>{val}</v></c>'
            return _re.sub(_pat(ref), new, xml, flags=_re.DOTALL)

        def sclr(xml, ref):
            new = f'<c r="{ref}"{_sty(ref)}/>' 
            return _re.sub(_pat(ref), new, xml, flags=_re.DOTALL)

        oggi_s = _excel_date(datetime.date.today())
        s1 = snum(s1, "B7", next_num)
        s1 = snum(s1, "B4", oggi_s)
        s1 = snum(s1, "B5", oggi_s)
        s1 = sstr(s1, "F3", nome_cli)
        s1 = sstr(s1, "F4", cli.get("indirizzo") or "")
        s1 = sstr(s1, "F5", cli.get("cap") or "")
        s1 = sstr(s1, "B11", marca_str)
        s1 = sstr(s1, "B12", targa_str)
        if km_str: s1 = snum(s1, "B13", km_str)
        if desc_lav: s1 = sstr(s1, "B16", desc_lav)

        for row in range(18, 44):
            i = row - 18
            if i < len(righe_fattura):
                v = righe_fattura[i]
                s1 = sstr(s1, f"B{row}", str(v["descrizione"]))
                s1 = snum(s1, f"E{row}", v["quantita"])
                s1 = snum(s1, f"F{row}", v["prezzo"])
                s1 = sform(s1, f"H{row}", f"E{row}*F{row}",
                           round(float(v["quantita"])*float(v["prezzo"]),2))
            else:
                s1 = sclr(s1, f"B{row}"); s1 = sclr(s1, f"E{row}")
                s1 = sclr(s1, f"F{row}")
                s1 = sform(s1, f"H{row}", f"E{row}*F{row}", 0)

        # Ricostruisce sharedStrings completamente (evita corruzione)
        new_entries = "".join(f"<si><t>{e}</t></si>" for e in str_list)
        new_ss = _re.sub(
            r"<sst[^>]*>.*?</sst>",
            f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
            f' count="{len(str_list)}" uniqueCount="{len(str_list)}">'
            f"{new_entries}</sst>",
            ss, flags=_re.DOTALL)

        tmpl["xl/worksheets/sheet1.xml"] = s1.encode("utf-8")
        tmpl["xl/sharedStrings.xml"]     = new_ss.encode("utf-8")

        # File da escludere (calcChain — Excel lo rigenera)
        escludi = {"xl/calcChain.xml"}

        with _zf.ZipFile(dest, "w", _zf.ZIP_DEFLATED) as zout:
            for name, data in tmpl.items():
                if name not in escludi:
                    zout.writestr(name, data)
    except Exception as ex:
        messagebox.showerror("Errore creazione fattura", str(ex)); return

    # ── Salva nel DB ──────────────────────────────────────────
    totale = round(totale_prev * (1 + IVA), 2)
    fid = db_commit(
        "INSERT INTO fatture (id_cliente,id_veicolo,data_fattura,totale,"
        "file_path,numero_fattura,iva_applicata) VALUES (%s,%s,%s,%s,%s,%s,TRUE)",
        (cid,vid,datetime.date.today(),totale,"fatture/"+nome_file,str(next_num)))
    for voce in righe_fattura:
        db_commit("INSERT INTO voci_fattura (id_fattura,descrizione,quantita,prezzo)"
                  " VALUES (%s,%s,%s,%s)",
                  (fid,voce["descrizione"],voce["quantita"],voce["prezzo"]))

    voci_testo = (desc_lav+"\n") if desc_lav else ""
    voci_testo += "\n".join("- "+v["descrizione"]+"  (Qt "+str(v["quantita"])+" x CHF "+f"{v['prezzo']:.2f}"+")"
                              for v in righe_fattura) if righe_fattura else "Fattura "+str(next_num)
    sid = db_commit(
        "INSERT INTO storico_interventi (id_veicolo,data_lavoro,descrizione,costo,note)"
        " VALUES (%s,%s,%s,%s,%s)",
        (vid,datetime.date.today(),voci_testo,totale,"[FATTURA:"+str(fid)+"]"))
    db_commit("UPDATE fatture SET id_storico=%s WHERE id_fattura=%s",(sid,fid))

    righe_fattura.clear(); aggiorna_tree_righe()
    desc_lav_txt.delete("1.0","end")
    aggiorna_lbl_num_fat()  # aggiorna il numero mostrato in cima

    messagebox.showinfo("Fattura creata",
        "Fattura N° "+str(next_num)+" salvata!")
    _avvia_watcher_fattura(dest, sid, vid, totale)
    try: apri_file(dest)
    except: pass



btn_fat = tk.Frame(tab_fatture); btn_fat.pack(fill="x",padx=10,pady=4)
tk.Button(btn_fat,text="➕ Aggiungi voce",command=aggiungi_riga, width=16).pack(side="left",padx=3)
tk.Button(btn_fat,text="🗑 Rimuovi voce", command=rimuovi_riga,  width=16).pack(side="left",padx=3)
tk.Button(btn_fat,text="🧾  GENERA FATTURA EXCEL",command=genera_fattura_excel,
          width=26,bg="#1565C0",fg="white",font=("Arial",10,"bold")).pack(side="left",padx=8)
lbl_totale = tk.Label(btn_fat,text="Totale: CHF 0.00",font=("Arial",10,"bold"))
lbl_totale.pack(side="right",padx=10)


# ══════════════════════════════════════════════════════════════
# SCHEDA 3 – STORICO VEICOLO
# ══════════════════════════════════════════════════════════════
tab_storico = tk.Frame(nb)
nb.add(tab_storico, text="🔧  Storico Veicolo")

top_sto = tk.Frame(tab_storico); top_sto.pack(fill="x",padx=10,pady=8)
tk.Label(top_sto,text="Cliente:").grid(row=0,column=0,padx=4,sticky="e")
sto_cli_var = tk.StringVar()
sto_cli_combo = ttk.Combobox(top_sto,textvariable=sto_cli_var,width=28)
sto_cli_combo.grid(row=0,column=1,padx=4)
tk.Label(top_sto,text="(scrivi per filtrare)",fg="gray",font=("Arial",8)).grid(row=1,column=0,columnspan=2,padx=4,sticky="w")
tk.Label(top_sto,text="Veicolo:").grid(row=0,column=2,padx=4,sticky="e")
sto_vei_var = tk.StringVar()
sto_vei_combo = ttk.Combobox(top_sto,textvariable=sto_vei_var,state="readonly",width=26)
sto_vei_combo.grid(row=0,column=3,padx=4)

sto_cli_dict = {}; sto_vei_dict = {}
sto_cli_tutti = []

def carica_sto_clienti():
    global sto_cli_tutti
    sto_cli_dict.clear(); sto_cli_tutti = []
    for r in db_fetch("SELECT id_cliente,cognome,nome FROM clienti ORDER BY cognome,nome"):
        k = f"{r['cognome']} {r['nome']}"
        sto_cli_dict[k] = r["id_cliente"]
        sto_cli_tutti.append(k)
    sto_cli_combo["values"] = sto_cli_tutti

def _sto_filtra_clienti(event=None):
    testo = sto_cli_var.get().lower()
    filtrati = [k for k in sto_cli_tutti if testo in k.lower()]
    sto_cli_combo["values"] = filtrati
    # se c'è una corrispondenza esatta carica subito
    if sto_cli_var.get() in sto_cli_dict:
        on_sto_cli_change()

def on_sto_cli_change(event=None):
    cid = sto_cli_dict.get(sto_cli_var.get())
    sto_vei_dict.clear(); sto_vei_combo.set(""); sto_vei_combo["values"] = []
    if not cid: return
    for r in db_fetch("SELECT id_veicolo,targa,marca,modello FROM veicoli WHERE id_cliente=%s",(cid,)):
        sto_vei_dict[f"{r['targa']} – {r['marca'] or ''} {r['modello'] or ''}".strip(" –")] = r["id_veicolo"]
    sto_vei_combo["values"] = list(sto_vei_dict.keys())
    if len(sto_vei_dict) == 1:
        sto_vei_combo.current(0); on_sto_vei_change()

def on_sto_vei_change(event=None):
    vid = sto_vei_dict.get(sto_vei_var.get())
    if vid: carica_storico_tree(vid)

sto_cli_combo.bind("<KeyRelease>", _sto_filtra_clienti)
sto_cli_combo.bind("<<ComboboxSelected>>", on_sto_cli_change)
sto_vei_combo.bind("<<ComboboxSelected>>", on_sto_vei_change)

lf_sto = tk.LabelFrame(tab_storico,text="Interventi  (doppio clic per leggere)")
lf_sto.pack(fill="both",expand=True,padx=10,pady=4)
cols_sto = ("data","descrizione","costo","note")
tree_sto = ttk.Treeview(lf_sto,columns=cols_sto,show="headings",height=16,style="Storico.Treeview")
for c,w,h in zip(cols_sto,(95,360,100,220),("Data","Descrizione","Costo CHF","Note")):
    tree_sto.heading(c,text=h); tree_sto.column(c,width=w)
sc_sto = ttk.Scrollbar(lf_sto,orient="vertical",command=tree_sto.yview)
tree_sto.config(yscrollcommand=sc_sto.set)
tree_sto.pack(side="left",fill="both",expand=True,padx=4,pady=4)
sc_sto.pack(side="right",fill="y")

def carica_storico_tree(vid):
    tree_sto.delete(*tree_sto.get_children())
    for r in db_fetch("SELECT id_intervento,data_lavoro,descrizione,costo,note FROM storico_interventi WHERE id_veicolo=%s ORDER BY data_lavoro DESC",(vid,)):
        # Nasconde il tag interno [FATTURA:XX] dalla colonna Note
        nota_visibile = r["note"] or ""
        import re as _re
        nota_visibile = _re.sub(r'\[FATTURA:\d+\]', '', nota_visibile).strip()
        tree_sto.insert("","end",iid=str(r["id_intervento"]),
            values=(str(r["data_lavoro"]),r["descrizione"],
                    f"{r['costo']:.2f}" if r["costo"] else "0.00", nota_visibile))

def leggi_intervento(event=None):
    """Doppio clic: apre popup grande in sola lettura."""
    sel = tree_sto.selection()
    if not sel: return
    iid = int(sel[0])
    r = db_fetch("SELECT * FROM storico_interventi WHERE id_intervento=%s",(iid,))[0]
    win = tk.Toplevel(root); win.title(f"Intervento del {r['data_lavoro']}"); win.geometry("700x520"); win.grab_set(); win.resizable(True,True)
    tk.Label(win,text=f"Data: {r['data_lavoro']}   |   Costo: CHF {r['costo'] or 0:.2f}",
             font=("Arial",11,"bold")).pack(pady=8)
    fr = tk.Frame(win); fr.pack(fill="both",expand=True,padx=10)
    tk.Label(fr,text="Descrizione:",font=("Arial",10,"bold")).grid(row=0,column=0,sticky="nw",pady=(0,2))
    t1 = tk.Text(fr,font=("Arial",10),wrap="word",height=10,state="normal")
    sb1 = ttk.Scrollbar(fr,orient="vertical",command=t1.yview); t1.config(yscrollcommand=sb1.set)
    t1.grid(row=1,column=0,sticky="nsew"); sb1.grid(row=1,column=1,sticky="ns")
    t1.insert("1.0",r["descrizione"]); t1.config(state="disabled")
    if r["note"]:
        tk.Label(fr,text="Note:",font=("Arial",10,"bold")).grid(row=2,column=0,sticky="nw",pady=(8,2))
        t2 = tk.Text(fr,font=("Arial",10),wrap="word",height=5,state="normal")
        sb2 = ttk.Scrollbar(fr,orient="vertical",command=t2.yview); t2.config(yscrollcommand=sb2.set)
        t2.grid(row=3,column=0,sticky="nsew"); sb2.grid(row=3,column=1,sticky="ns")
        t2.insert("1.0",r["note"]); t2.config(state="disabled")
        fr.rowconfigure(3,weight=1)
    fr.rowconfigure(1,weight=2); fr.columnconfigure(0,weight=1)
    tk.Button(win,text="✏️ Modifica",command=lambda:[win.destroy(), modifica_intervento()],
              width=14).pack(side="left",padx=20,pady=10)
    tk.Button(win,text="Chiudi",command=win.destroy,width=12).pack(side="right",padx=20,pady=10)

tree_sto.bind("<Double-1>", leggi_intervento)

def _popup_intervento(win_title, dati_esistenti=None):
    win = tk.Toplevel(root); win.title(win_title); win.geometry("750x600"); win.grab_set(); win.resizable(True,True)
    tk.Label(win,text="Data (AAAA-MM-GG):",font=("Arial",10)).grid(row=0,column=0,padx=12,pady=10,sticky="e")
    e_data = tk.Entry(win,width=16,font=("Arial",10))
    e_data.insert(0,str(dati_esistenti["data_lavoro"]) if dati_esistenti else datetime.date.today().isoformat())
    e_data.grid(row=0,column=1,padx=12,pady=10,sticky="w")
    tk.Label(win,text="Costo CHF:",font=("Arial",10)).grid(row=0,column=2,padx=12,pady=10,sticky="e")
    e_costo = tk.Entry(win,width=12,font=("Arial",10))
    e_costo.insert(0,str(dati_esistenti["costo"] or 0) if dati_esistenti else "0")
    e_costo.grid(row=0,column=3,padx=12,pady=10,sticky="w")
    tk.Label(win,text="Descrizione*:",font=("Arial",10,"bold")).grid(row=1,column=0,padx=12,pady=6,sticky="ne")
    e_desc = tk.Text(win,font=("Arial",10),wrap="word")
    sb1 = ttk.Scrollbar(win,orient="vertical",command=e_desc.yview); e_desc.config(yscrollcommand=sb1.set)
    e_desc.grid(row=1,column=1,columnspan=3,padx=(12,0),pady=6,sticky="nsew"); sb1.grid(row=1,column=4,padx=(0,12),pady=6,sticky="ns")
    if dati_esistenti: e_desc.insert("1.0",dati_esistenti["descrizione"])
    e_desc.focus()
    tk.Label(win,text="Note:",font=("Arial",10)).grid(row=2,column=0,padx=12,pady=6,sticky="ne")
    e_note = tk.Text(win,font=("Arial",10),wrap="word")
    sb2 = ttk.Scrollbar(win,orient="vertical",command=e_note.yview); e_note.config(yscrollcommand=sb2.set)
    e_note.grid(row=2,column=1,columnspan=3,padx=(12,0),pady=6,sticky="nsew"); sb2.grid(row=2,column=4,padx=(0,12),pady=6,sticky="ns")
    if dati_esistenti: e_note.insert("1.0",dati_esistenti["note"] or "")
    win.columnconfigure(1,weight=1); win.rowconfigure(1,weight=3); win.rowconfigure(2,weight=1)
    return win, e_data, e_desc, e_costo, e_note

def aggiungi_intervento():
    vid = sto_vei_dict.get(sto_vei_var.get())
    if not vid:
        messagebox.showwarning("Attenzione","Seleziona prima cliente e veicolo"); return
    win, e_data, e_desc, e_costo, e_note = _popup_intervento("Nuovo Intervento")
    def salva():
        desc = e_desc.get("1.0","end").strip()
        if not desc:
            messagebox.showwarning("Errore","La descrizione è obbligatoria",parent=win); return
        try: costo = float(e_costo.get().replace(",","."))
        except: costo = 0
        db_commit("INSERT INTO storico_interventi (id_veicolo,data_lavoro,descrizione,costo,note) VALUES (%s,%s,%s,%s,%s)",
            (vid,e_data.get().strip(),desc,costo,e_note.get("1.0","end").strip() or None))
        messagebox.showinfo("OK","Intervento salvato!",parent=win)
        win.destroy(); carica_storico_tree(vid)
    tk.Button(win,text="💾  Salva Intervento",command=salva,width=22,bg="#4CAF50",fg="white",font=("Arial",11,"bold")).grid(row=3,column=0,columnspan=5,pady=14)

def modifica_intervento():
    sel = tree_sto.selection()
    if not sel:
        messagebox.showwarning("Attenzione","Seleziona un intervento"); return
    iid = int(sel[0])
    r = db_fetch("SELECT * FROM storico_interventi WHERE id_intervento=%s",(iid,))[0]
    win, e_data, e_desc, e_costo, e_note = _popup_intervento("Modifica Intervento", dati_esistenti=r)
    def salva():
        desc = e_desc.get("1.0","end").strip()
        if not desc:
            messagebox.showwarning("Errore","Descrizione obbligatoria",parent=win); return
        try: costo = float(e_costo.get().replace(",","."))
        except: costo = 0
        db_commit("UPDATE storico_interventi SET data_lavoro=%s,descrizione=%s,costo=%s,note=%s WHERE id_intervento=%s",
            (e_data.get().strip(),desc,costo,e_note.get("1.0","end").strip() or None,iid))
        messagebox.showinfo("OK","Aggiornato!",parent=win)
        win.destroy()
        vid = sto_vei_dict.get(sto_vei_var.get())
        if vid: carica_storico_tree(vid)
    tk.Button(win,text="💾  Salva Modifiche",command=salva,width=22,bg="#4CAF50",fg="white",font=("Arial",11,"bold")).grid(row=3,column=0,columnspan=5,pady=14)

def elimina_intervento():
    sel = tree_sto.selection()
    if not sel:
        messagebox.showwarning("Attenzione","Seleziona un intervento"); return
    if not messagebox.askyesno("Conferma","Eliminare questo intervento?"): return
    db_commit("DELETE FROM storico_interventi WHERE id_intervento=%s",(int(sel[0]),))
    vid = sto_vei_dict.get(sto_vei_var.get())
    if vid: carica_storico_tree(vid)

btn_sto = tk.Frame(tab_storico); btn_sto.pack(fill="x",padx=10,pady=9)
tk.Button(btn_sto,text="➕ Nuovo Intervento",command=aggiungi_intervento,width=30,font=("Arial",11,"bold")).pack(side="left",padx=10)
tk.Button(btn_sto,text="✏️ Modifica",         command=modifica_intervento,width=30,font=("Arial",11,"bold")).pack(side="left",padx=10)
tk.Button(btn_sto,text="🗑 Elimina",          command=elimina_intervento, width=30,font=("Arial",11,"bold")).pack(side="left",padx=10)


# ══════════════════════════════════════════════════════════════
# SCHEDA 4 – MAGAZZINO GOMME
# ══════════════════════════════════════════════════════════════
tab_gomme = tk.Frame(nb)
nb.add(tab_gomme, text="🛞  Magazzino Gomme")

top_gom = tk.Frame(tab_gomme); top_gom.pack(fill="x",padx=10,pady=6)
tk.Label(top_gom,text="Cliente:").pack(side="left",padx=4)
gom_cli_var = tk.StringVar()
gom_cli_combo = ttk.Combobox(top_gom,textvariable=gom_cli_var,width=22)
gom_cli_combo.pack(side="left",padx=4)
tk.Label(top_gom,text="(scrivi per filtrare)",fg="gray",font=("Arial",8)).pack(side="left",padx=(0,6))

tk.Label(top_gom,text="Targa:").pack(side="left",padx=(8,4))
gom_targa_var = tk.StringVar()
tk.Entry(top_gom, textvariable=gom_targa_var, width=12).pack(side="left",padx=4)
tk.Label(top_gom,text="(filtra per targa)",fg="gray",font=("Arial",8)).pack(side="left",padx=(0,6))
gom_targa_var.trace_add("write", lambda *_: carica_gomme())

tk.Label(top_gom,text="Stagione:").pack(side="left",padx=(8,4))
gom_stag_var = tk.StringVar(value="Tutti")
ttk.Combobox(top_gom,textvariable=gom_stag_var,state="readonly",width=12,
    values=["Tutti","Estiva","Invernale","All-Season","Altri"]).pack(side="left",padx=4)
tk.Button(top_gom,text="🔄 Aggiorna",command=lambda: carica_gomme()).pack(side="left",padx=8)
tk.Button(top_gom,text="🖨 Stampa Magazzino",command=lambda: stampa_magazzino_gomme()).pack(side="left",padx=4)

lf_gom = tk.LabelFrame(tab_gomme,text="Gomme  –  doppio clic su 'Montate' per cambiarlo  |  doppio clic su 'Dep. Pagato' per cambiarlo  |  tasto Modifica per tutto il resto")
lf_gom.pack(fill="both",expand=True,padx=10,pady=4)
cols_gom = ("montate","cliente","targa_vei","marca","modello","stagione","misura","qty","stato","posizione","dep")
tree_gom = ttk.Treeview(lf_gom,columns=cols_gom,show="headings",height=16)
for c,w,h in zip(cols_gom,(70,120,80,85,95,85,80,40,95,100,90),
                 ("Montate","Cliente","Targa","Marca","Modello","Stagione","Misura","Qt","Stato","Posizione","Dep. Pagato")):
    tree_gom.heading(c,text=h); tree_gom.column(c,width=w)
tree_gom.column("montate", anchor="center")
sc_gom = ttk.Scrollbar(lf_gom,orient="vertical",command=tree_gom.yview)
tree_gom.config(yscrollcommand=sc_gom.set)
tree_gom.pack(side="left",fill="both",expand=True,padx=4,pady=4)
sc_gom.pack(side="right",fill="y")

# tag colori per la colonna Montate
tree_gom.tag_configure("montate_si",  background="#c8f7c5", foreground="#1a7a16")  # verde
tree_gom.tag_configure("montate_no",  background="#ffffff", foreground="#888888")  # grigio

gom_cli_dict2 = {}
gom_cli_tutti2 = []

def carica_gom_clienti():
    global gom_cli_tutti2
    gom_cli_dict2.clear(); gom_cli_tutti2 = []
    for k, v in [("Tutti","all"), ("— Officina (proprie) —", None)]:
        gom_cli_dict2[k] = v; gom_cli_tutti2.append(k)
    for r in db_fetch("SELECT id_cliente,cognome,nome FROM clienti ORDER BY cognome,nome"):
        k = f"{r['cognome']} {r['nome']}"
        gom_cli_dict2[k] = r["id_cliente"]; gom_cli_tutti2.append(k)
    gom_cli_combo["values"] = gom_cli_tutti2
    if not gom_cli_var.get(): gom_cli_combo.set("Tutti")

def _gom_filtra_clienti(event=None):
    testo = gom_cli_var.get().lower()
    filtrati = [k for k in gom_cli_tutti2 if testo in k.lower()]
    gom_cli_combo["values"] = filtrati
    # Aggiorna la tabella in tempo reale (come per la targa)
    carica_gomme()

gom_cli_combo.bind("<KeyRelease>", _gom_filtra_clienti)
gom_cli_combo.bind("<<ComboboxSelected>>", lambda e: carica_gomme())
gom_stag_var.trace_add("write", lambda *_: carica_gomme())

def carica_gomme():
    tree_gom.delete(*tree_gom.get_children())
    testo_cli  = gom_cli_var.get()
    cli_sel    = gom_cli_dict2.get(testo_cli, "partial")  # "partial" = testo libero
    stag       = gom_stag_var.get()
    targa_filt = gom_targa_var.get().strip().upper()
    sql = """SELECT g.id_gomma,
               IFNULL(CONCAT(c.cognome,' ',c.nome),'— Officina —') AS cliente,
               IFNULL(g.targa_veicolo,'') AS targa_vei,
               g.marca, g.modello, g.stagione, g.misura, g.quantita,
               g.stato, g.posizione, g.deposito_pagato,
               IFNULL(g.montate,0) AS montate
             FROM magazzino_gomme g
             LEFT JOIN clienti c ON c.id_cliente=g.id_cliente
             WHERE 1=1"""
    params = []
    if cli_sel == "all" or not testo_cli:
        pass  # mostra tutto
    elif cli_sel is None:
        sql += " AND g.id_cliente IS NULL"
    elif cli_sel == "partial":
        # testo libero: filtra per somiglianza nome cliente
        if testo_cli.lower() not in ("tutti",""):
            sql += " AND CONCAT(IFNULL(c.cognome,''),' ',IFNULL(c.nome,'')) LIKE %s"
            params.append(f"%{testo_cli}%")
    else:
        sql += " AND g.id_cliente=%s"
        params.append(cli_sel)
    if stag != "Tutti":
        sql += " AND g.stagione=%s"
        params.append(stag)
    if targa_filt:
        sql += " AND g.targa_veicolo LIKE %s"
        params.append(f"%{targa_filt}%")
    sql += " ORDER BY cliente, g.targa_veicolo, g.montate DESC, g.marca"
    palette = [
        {"si": "#b5d8f0", "no": "#e8f0f8"},   # blu chiaro per tutti i clienti
        {"si": "#b5d8f0", "no": "#e8f0f8"},
        {"si": "#b5d8f0", "no": "#e8f0f8"},
        {"si": "#b5d8f0", "no": "#e8f0f8"},
    ]
    prev_cliente = None
    color_idx    = -1
    for r in db_fetch(sql, tuple(params)):
        cliente = r["cliente"]
        if cliente != prev_cliente:
            if prev_cliente is not None:
                tree_gom.insert("", "end", iid=f"sep_{r['id_gomma']}",
                    tags=("separatore",),
                    values=("","","","","","","","","","",""))
            prev_cliente = cliente
            color_idx = (color_idx + 1) % len(palette)
        montate  = bool(r["montate"])
        mont_txt = "🟢 Sì" if montate else "○ No"
        dep      = "✅ Sì" if r["deposito_pagato"] else "❌ No"
        tag      = f"cli_{color_idx}_{'si' if montate else 'no'}"
        tree_gom.insert("", "end", iid=str(r["id_gomma"]), tags=(tag,),
            values=(mont_txt, cliente, r["targa_vei"], r["marca"], r["modello"] or "",
                    r["stagione"] or "", r["misura"] or "", r["quantita"],
                    r["stato"] or "", r["posizione"] or "", dep))
    for i, pal in enumerate(palette):
        tree_gom.tag_configure(f"cli_{i}_si",
            background=pal["si"], foreground="#1a7a16", font=("Arial",10,"bold"))
        tree_gom.tag_configure(f"cli_{i}_no",
            background=pal["no"], foreground="#444444")
    tree_gom.tag_configure("separatore", background="#cccccc")

def toggle_montate(gid):
    """Alterna il flag Montate su una gomma specifica (usato dal pulsante dedicato)."""
    rows = db_fetch("SELECT montate FROM magazzino_gomme WHERE id_gomma=%s",(gid,))
    if not rows: return
    attuale = bool(rows[0]["montate"])
    db_commit("UPDATE magazzino_gomme SET montate=%s WHERE id_gomma=%s",(0 if attuale else 1, gid))
    carica_gomme()
    try: tree_gom.selection_set(str(gid)); tree_gom.see(str(gid))
    except: pass


def stampa_magazzino_gomme():
    cli_sel = gom_cli_dict2.get(gom_cli_var.get(), "all")
    stag    = gom_stag_var.get()
    sql = """SELECT IFNULL(CONCAT(c.cognome,' ',c.nome),'— Officina —') AS cliente,
               g.marca, g.modello, g.stagione, g.misura, g.quantita,
               g.stato, g.posizione, g.deposito_pagato, IFNULL(g.montate,0) AS montate
             FROM magazzino_gomme g
             LEFT JOIN clienti c ON c.id_cliente=g.id_cliente
             WHERE 1=1"""
    params = []
    if cli_sel != "all":
        if cli_sel is None:
            sql += " AND g.id_cliente IS NULL"
        else:
            sql += " AND g.id_cliente=%s"
            params.append(cli_sel)
    if stag != "Tutti":
        sql += " AND g.stagione=%s"
        params.append(stag)
    sql += " ORDER BY cliente, g.montate DESC, g.marca"
    rows = db_fetch(sql, tuple(params))
    if not rows:
        messagebox.showinfo("Stampa","Nessuna gomma da stampare."); return

    # Genera PDF A4 verticale con font piccolo per stare su una riga
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        import tempfile, os as _os

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        tmp_path = tmp.name; tmp.close()

        doc = SimpleDocTemplate(tmp_path, pagesize=A4,
            leftMargin=8*mm, rightMargin=8*mm, topMargin=10*mm, bottomMargin=10*mm)

        styles = getSampleStyleSheet()
        small  = ParagraphStyle("small", fontSize=7, leading=9)
        header_s = ParagraphStyle("hdr", fontSize=9, leading=11, fontName="Helvetica-Bold")

        titolo_txt = f"MAGAZZINO GOMME – Garage Tito   |   {datetime.date.today().strftime('%d/%m/%Y')}"
        if gom_cli_var.get() and gom_cli_var.get() not in ("Tutti",""):
            titolo_txt += f"   |   Cliente: {gom_cli_var.get()}"

        story = [Paragraph(titolo_txt, header_s), Spacer(1, 3*mm)]

        # Intestazione tabella
        col_hdrs = ["Mont.", "Cliente", "Marca", "Modello", "Stagione", "Misura", "Qt", "Stato", "Posizione", "Dep."]
        data = [col_hdrs]
        for r in rows:
            data.append([
                "SI" if r["montate"] else "no",
                r["cliente"],
                r["marca"],
                r["modello"] or "",
                r["stagione"] or "",
                r["misura"] or "",
                str(r["quantita"]),
                r["stato"] or "",
                r["posizione"] or "",
                "Si" if r["deposito_pagato"] else "No",
            ])

        # Larghezze colonne proporzionali all'A4 (larghezza utile ~194mm)
        col_w = [10*mm, 32*mm, 22*mm, 26*mm, 16*mm, 18*mm, 8*mm, 20*mm, 22*mm, 10*mm]

        t = Table(data, colWidths=col_w, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,0),  colors.HexColor("#1565C0")),
            ("TEXTCOLOR",    (0,0), (-1,0),  colors.white),
            ("FONTNAME",     (0,0), (-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 7),
            ("LEADING",      (0,0), (-1,-1), 9),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.HexColor("#e8f0f8"), colors.white]),
            ("GRID",         (0,0), (-1,-1), 0.3, colors.grey),
            ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
            ("LEFTPADDING",  (0,0), (-1,-1), 2),
            ("RIGHTPADDING", (0,0), (-1,-1), 2),
            ("TOPPADDING",   (0,0), (-1,-1), 1),
            ("BOTTOMPADDING",(0,0), (-1,-1), 1),
        ]))
        story.append(t)
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph(f"Totale: {len(rows)} voci", small))

        doc.build(story)
        apri_file(tmp_path)

    except ImportError:
        # Fallback testo se reportlab non è installato
        messagebox.showwarning("Info",
            "Installa reportlab per la stampa PDF A4:\n  pip install reportlab\n\n"
            "Nel frattempo uso la stampa testo.")
        W_CLI=20; W_MAR=14; W_MOD=14; W_MIS=12; W_STA=12; W_POS=12
        sep = "─"*100
        header = f"{'Mon':<5} {'Cliente':<{W_CLI}} {'Marca':<{W_MAR}} {'Modello':<{W_MOD}} {'Stag.':<10} {'Misura':<{W_MIS}} {'Qt':<4} {'Stato':<{W_STA}} {'Pos.':<{W_POS}} Dep."
        lines = [f"{'MAGAZZINO GOMME – Garage Tito':^100}", sep, header, sep]
        for r in rows:
            dep = "Si" if r["deposito_pagato"] else "No"
            mont = "SI" if r["montate"] else "  "
            lines.append(f"{mont:<5} {r['cliente']:<{W_CLI}} {r['marca']:<{W_MAR}} {r['modello'] or '':<{W_MOD}} "
                         f"{r['stagione'] or '':<10} {r['misura'] or '':<{W_MIS}} {r['quantita']:<4} "
                         f"{r['stato'] or '':<{W_STA}} {r['posizione'] or '':<{W_POS}} {dep}")
        lines += [sep, f"Totale: {len(rows)}   Stampato il {datetime.date.today().strftime('%d/%m/%Y')}"]
        stampa_testo("Magazzino Gomme", "\n".join(lines))

def finestra_aggiungi_gomma():
    """
    Finestra per aggiungere gomme al magazzino.
    Permette di aggiungere più set (es. estive + invernali) per lo stesso cliente
    in una sola volta, con la spunta 'Montate' su uno solo o entrambi.
    """
    win = tk.Toplevel(root)
    win.title("Aggiungi Gomme al Magazzino")
    win.resizable(True, True)
    win.grab_set()

    # ── Cliente + Targa ──────────────────────────────────────
    top = tk.Frame(win); top.pack(fill="x", padx=12, pady=8)
    tk.Label(top, text="Cliente:", font=("Arial",10,"bold")).pack(side="left")

    cli_var2   = tk.StringVar()
    cli_cb2    = ttk.Combobox(top, textvariable=cli_var2, width=26)
    cli_ids2   = {"— Officina (proprie) —": None}
    cli_tutti2 = ["— Officina (proprie) —"]
    for r in db_fetch("SELECT id_cliente,cognome,nome,tipo_cliente,ragione_sociale FROM clienti ORDER BY cognome,nome"):
        if r.get("tipo_cliente") == "azienda":
            k = r.get("ragione_sociale") or r["cognome"]
        else:
            k = f"{r['cognome']} {r['nome']}"
        cli_ids2[k] = r["id_cliente"]
        cli_tutti2.append(k)
    # Dizionario targa→(id_cliente, nome) per ricerca inversa
    targa_a_cliente2 = {}
    for r in db_fetch("SELECT v.targa, c.id_cliente, c.cognome, c.nome, c.tipo_cliente, c.ragione_sociale "
                      "FROM veicoli v JOIN clienti c ON c.id_cliente=v.id_cliente WHERE v.targa IS NOT NULL"):
        if r.get("tipo_cliente") == "azienda":
            nome = r.get("ragione_sociale") or r["cognome"]
        else:
            nome = f"{r['cognome']} {r['nome']}"
        targa_a_cliente2[r["targa"].upper()] = (r["id_cliente"], nome)
    all_targhe2 = sorted(targa_a_cliente2.keys())

    cli_cb2["values"] = cli_tutti2
    cli_cb2.set("— Officina (proprie) —")
    cli_cb2.pack(side="left", padx=8)
    tk.Label(top, text="(scrivi per filtrare)", fg="gray", font=("Arial",8)).pack(side="left")

    tk.Label(top, text="  Targa:", font=("Arial",10,"bold")).pack(side="left", padx=(10,4))
    targa_var2 = tk.StringVar()
    targa_cb2  = ttk.Combobox(top, textvariable=targa_var2, width=14)
    targa_cb2["values"] = all_targhe2
    targa_cb2.pack(side="left", padx=4)
    tk.Label(top, text="(scrivi targa)", fg="gray", font=("Arial",8)).pack(side="left")
    tutte_targhe2 = []

    def _aggiorna_targhe_per_cliente(cid2):
        nonlocal tutte_targhe2
        if cid2 is None:
            tutte_targhe2 = []
            targa_cb2["values"] = ["(nessuna)"]
            targa_cb2.set("(nessuna)")
            return
        veicoli = db_fetch("SELECT targa FROM veicoli WHERE id_cliente=%s ORDER BY targa", (cid2,))
        targhe = [v["targa"] for v in veicoli if v["targa"]]
        tutte_targhe2 = targhe
        targa_cb2["values"] = targhe if targhe else ["(nessuna)"]
        targa_cb2.set(targhe[0] if targhe else "(nessuna)")

    def on_cli_change2(event=None):
        cid2 = cli_ids2.get(cli_var2.get())
        _aggiorna_targhe_per_cliente(cid2)

    def _filtra_clienti2(event=None):
        testo = cli_var2.get().lower()
        filtrati = [k for k in cli_tutti2 if testo in k.lower()]
        cli_cb2["values"] = filtrati

    def _filtra_targa2(event=None):
        testo = targa_var2.get().strip().upper()
        if not testo:
            targa_cb2["values"] = tutte_targhe2 if tutte_targhe2 else all_targhe2
            return
        matches = [t for t in all_targhe2 if testo in t]
        targa_cb2["values"] = matches if matches else all_targhe2

    def on_targa_selected2(event=None):
        targa = targa_var2.get().strip().upper()
        if targa in targa_a_cliente2:
            cid2, nome_cli = targa_a_cliente2[targa]
            cli_var2.set(nome_cli)
            _aggiorna_targhe_per_cliente(cid2)
            targa_var2.set(targa)

    cli_cb2.bind("<KeyRelease>", _filtra_clienti2)
    cli_cb2.bind("<<ComboboxSelected>>", on_cli_change2)
    targa_cb2.bind("<KeyRelease>", _filtra_targa2)
    targa_cb2.bind("<<ComboboxSelected>>", on_targa_selected2)


    # ── Area set di gomme (scrollabile) ──────────────────────
    lf_sets = tk.LabelFrame(win, text="Set di Gomme  (puoi aggiungere più set per lo stesso cliente)")
    lf_sets.pack(fill="both", expand=True, padx=12, pady=4)

    canvas_s = tk.Canvas(lf_sets, highlightthickness=0)
    sb_s = ttk.Scrollbar(lf_sets, orient="vertical", command=canvas_s.yview)
    canvas_s.configure(yscrollcommand=sb_s.set)
    canvas_s.pack(side="left", fill="both", expand=True)
    sb_s.pack(side="right", fill="y")

    inner = tk.Frame(canvas_s)
    inner_win_id = canvas_s.create_window((0,0), window=inner, anchor="nw")

    def on_inner_configure(event):
        canvas_s.configure(scrollregion=canvas_s.bbox("all"))
        canvas_s.itemconfig(inner_win_id, width=canvas_s.winfo_width())
    inner.bind("<Configure>", on_inner_configure)
    canvas_s.bind("<Configure>", lambda e: canvas_s.itemconfig(inner_win_id, width=e.width))

    sets_data = []   # lista di dict con i widget di ogni set

    def aggiungi_set(dati_default=None):
        """Aggiunge un pannello per un nuovo set di gomme."""
        idx = len(sets_data)
        color_bg = "#f0f8ff" if idx % 2 == 0 else "#f8fff0"

        frm = tk.LabelFrame(inner, text=f"Set {idx+1}",
                            bg=color_bg, padx=8, pady=6)
        frm.pack(fill="x", padx=6, pady=4)

        def campo(r, lbl, key, default="", width=20):
            tk.Label(frm, text=lbl, bg=color_bg).grid(row=r, column=0, sticky="e", padx=6, pady=3)
            e = tk.Entry(frm, width=width)
            e.insert(0, default)
            e.grid(row=r, column=1, sticky="w", padx=6, pady=3)
            return e

        e_marca   = campo(0, "Marca*",              "marca",   dati_default.get("marca","")   if dati_default else "")
        e_modello = campo(1, "Modello",              "modello", dati_default.get("modello","") if dati_default else "")
        e_misura  = campo(2, "Misura (es.205/55R16)","misura",  dati_default.get("misura","")  if dati_default else "")
        e_qty     = campo(3, "Quantità",             "quantita","4")
        e_pos     = campo(4, "Posizione scaffale",   "pos",     "")

        tk.Label(frm, text="Stagione:", bg=color_bg).grid(row=0, column=2, sticky="e", padx=(16,4), pady=3)
        stag_v = tk.StringVar(value="Estiva")
        ttk.Combobox(frm, textvariable=stag_v, state="readonly", width=12,
            values=["Estiva","Invernale","All-Season","Altri"]).grid(row=0, column=3, sticky="w", padx=4, pady=3)

        tk.Label(frm, text="Stato:", bg=color_bg).grid(row=1, column=2, sticky="e", padx=(16,4), pady=3)
        stato_v = tk.StringVar(value="Nuovo")
        ttk.Combobox(frm, textvariable=stato_v, state="readonly", width=12,
            values=["Nuovo","Usato buono","Usato","Da smaltire"]).grid(row=1, column=3, sticky="w", padx=4, pady=3)

        dep_v  = tk.BooleanVar(value=False)
        mont_v = tk.BooleanVar(value=False)
        tk.Checkbutton(frm, text="Deposito pagato", variable=dep_v,
                       bg=color_bg).grid(row=2, column=2, columnspan=2, sticky="w", padx=4, pady=2)
        tk.Checkbutton(frm, text="🟢 MONTATE sul veicolo", variable=mont_v,
                       bg=color_bg, font=("Arial",10,"bold")).grid(row=3, column=2, columnspan=2, sticky="w", padx=4, pady=2)

        # Note — riga 5, separata dalla posizione
        tk.Label(frm, text="Note:", bg=color_bg).grid(row=5, column=0, sticky="ne", padx=6, pady=3)
        e_note = tk.Text(frm, width=22, height=3, bg="white", wrap="word")
        e_note.grid(row=5, column=1, columnspan=3, sticky="ew", padx=6, pady=3)

        # Pulsante rimuovi: in alto a destra nel titolo del frame
        if idx > 0:
            def rimuovi(f=frm, s=sets_data):
                to_remove = None
                for d in s:
                    if d["frame"] is f:
                        to_remove = d; break
                if to_remove:
                    s.remove(to_remove)
                    f.destroy()
                    for i, sd in enumerate(s):
                        sd["frame"].config(text=f"Set {i+1}")
                    win.update_idletasks()
                    canvas_s.configure(scrollregion=canvas_s.bbox("all"))
            tk.Button(frm, text="✕ Rimuovi", command=rimuovi,
                      fg="white", bg="#e53935", relief="flat",
                      font=("Arial",8,"bold"), padx=6).grid(
                      row=0, column=4, rowspan=2, padx=(12,4), pady=4, sticky="ne")

        sets_data.append({
            "frame": frm,
            "marca": e_marca, "modello": e_modello, "misura": e_misura,
            "quantita": e_qty, "posizione": e_pos,
            "stagione": stag_v, "stato": stato_v,
            "deposito": dep_v, "montate": mont_v,
            "note": e_note
        })
        # aggiusta altezza canvas
        win.update_idletasks()
        canvas_s.configure(scrollregion=canvas_s.bbox("all"))

    # Aggiunge il primo set di default
    aggiungi_set()

    # ── Pulsanti in fondo ─────────────────────────────────────
    btn_row = tk.Frame(win); btn_row.pack(fill="x", padx=12, pady=8)

    tk.Button(btn_row, text="➕ Aggiungi un altro set di gomme",
              command=aggiungi_set, width=30,
              bg="#FF9800", fg="white", font=("Arial",10,"bold")).pack(side="left", padx=4)

    def salva_tutto():
        try:
            sel_cliente = cli_var2.get()
            cid = cli_ids2.get(sel_cliente)
            targa_sel = targa_var2.get()
            targa_save = targa_sel if targa_sel and targa_sel != "(nessuna)" else None

            if not sets_data:
                messagebox.showwarning("Attenzione", "Nessun set da salvare.", parent=win)
                return

            salvati = 0
            errori  = []

            for i, s in enumerate(sets_data):
                try:
                    marca = s["marca"].get().strip()
                    if not marca:
                        errori.append(f"Set {i+1}: la marca è obbligatoria")
                        continue
                    try:
                        qty = int(s["quantita"].get().strip())
                    except Exception:
                        qty = 4

                    modello  = s["modello"].get().strip()   or None
                    misura   = s["misura"].get().strip()    or None
                    posiz    = s["posizione"].get().strip() or None
                    stagione = s["stagione"].get()
                    stato    = s["stato"].get()
                    dep      = 1 if s["deposito"].get() else 0
                    mont     = 1 if s["montate"].get() else 0
                    note_txt = s["note"].get("1.0","end").strip() or None

                    db_commit(
                        "INSERT INTO magazzino_gomme "
                        "(id_cliente, targa_veicolo, marca, modello, stagione, misura, "
                        " quantita, stato, posizione, deposito_pagato, montate, note) "
                        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                        (cid, targa_save, marca, modello, stagione, misura,
                         qty, stato, posiz, dep, mont, note_txt))
                    salvati += 1

                except Exception as ex_set:
                    errori.append(f"Set {i+1}: errore – {ex_set}")

            if errori and salvati == 0:
                messagebox.showerror("Errore",
                    "Nessuna gomma salvata.\n\n" + "\n".join(errori), parent=win)
            elif errori:
                messagebox.showwarning("Parzialmente salvato",
                    f"Salvati {salvati} set.\nErrori:\n" + "\n".join(errori), parent=win)
                carica_gomme()
            else:
                messagebox.showinfo("✅ Salvato",
                    f"{salvati} set di gomme aggiunti al magazzino!", parent=win)
                win.destroy()
                carica_gomme()

        except Exception as ex:
            messagebox.showerror("Errore imprevisto",
                f"Errore durante il salvataggio:\n{ex}\n\n"
                "Verifica che il database MySQL sia avviato.", parent=win)

    tk.Button(btn_row, text="💾 Salva tutti nel Magazzino",
              command=salva_tutto, width=28,
              bg="#4CAF50", fg="white", font=("Arial",10,"bold")).pack(side="right", padx=4)

    win.update_idletasks()
    win.geometry("680x560")

def modifica_gomma(gid_override=None):
    if gid_override:
        gid = gid_override
    else:
        sel = tree_gom.selection()
        if not sel:
            messagebox.showwarning("Attenzione","Seleziona una riga"); return
        if sel[0].startswith("sep_"):
            messagebox.showwarning("Attenzione","Seleziona una riga gomma, non il separatore"); return
        gid = int(sel[0])

    r = db_fetch("SELECT * FROM magazzino_gomme WHERE id_gomma=%s",(gid,))[0]

    win = tk.Toplevel(root)
    win.title(f"Modifica Gomma – {r['marca'] or ''} {r['modello'] or ''}".strip())
    win.geometry("500x560"); win.grab_set(); win.resizable(True, True)

    # Griglia campi a due colonne
    def lbl(row, col, testo):
        tk.Label(win, text=testo, anchor="e").grid(row=row, column=col, padx=10, pady=5, sticky="e")

    def entry(row, col, valore="", width=22):
        e = tk.Entry(win, width=width)
        e.insert(0, str(valore) if valore else "")
        e.grid(row=row, column=col, padx=10, pady=5, sticky="w")
        return e

    lbl(0, 0, "Marca*:");        e_marca   = entry(0, 1, r["marca"] or "")
    lbl(1, 0, "Modello:");       e_modello = entry(1, 1, r["modello"] or "")
    lbl(2, 0, "Misura:");        e_misura  = entry(2, 1, r["misura"] or "")
    lbl(3, 0, "Quantità:");      e_qty     = entry(3, 1, r["quantita"] or 4, width=10)
    lbl(4, 0, "Posizione:");     e_pos     = entry(4, 1, r["posizione"] or "")

    lbl(0, 2, "Stagione:");
    stag_m = tk.StringVar(value=r["stagione"] or "Estiva")
    ttk.Combobox(win, textvariable=stag_m, state="readonly", width=14,
        values=["Estiva","Invernale","All-Season","Altri"]).grid(row=0, column=3, padx=10, pady=5, sticky="w")

    lbl(1, 2, "Stato:");
    stato_m = tk.StringVar(value=r["stato"] or "Nuovo")
    ttk.Combobox(win, textvariable=stato_m, state="readonly", width=14,
        values=["Nuovo","Usato buono","Usato","Da smaltire"]).grid(row=1, column=3, padx=10, pady=5, sticky="w")

    lbl(2, 2, "Targa veicolo:");
    e_targa_m = entry(2, 3, r.get("targa_veicolo") or "", width=14)

    dep_m  = tk.BooleanVar(value=bool(r["deposito_pagato"]))
    mont_m = tk.BooleanVar(value=bool(r.get("montate",0)))
    tk.Checkbutton(win, text="Deposito pagato", variable=dep_m).grid(
        row=3, column=2, columnspan=2, sticky="w", padx=10, pady=2)
    tk.Checkbutton(win, text="🟢 Gomme MONTATE sul veicolo", variable=mont_m,
                   font=("Arial",10,"bold")).grid(
        row=4, column=2, columnspan=2, sticky="w", padx=10, pady=2)

    # Note — occupa tutta la larghezza
    tk.Label(win, text="Note:", anchor="e").grid(row=5, column=0, padx=10, pady=5, sticky="ne")
    e_note_m = tk.Text(win, width=44, height=5, wrap="word")
    e_note_m.insert("1.0", r["note"] or "")
    e_note_m.grid(row=5, column=1, columnspan=3, padx=10, pady=5, sticky="ew")
    win.columnconfigure(1, weight=1); win.columnconfigure(3, weight=1)

    def salva_m():
        marca = e_marca.get().strip()
        if not marca:
            messagebox.showwarning("Errore","La marca è obbligatoria",parent=win); return
        try: qty = int(e_qty.get())
        except: qty = 1
        targa_m = e_targa_m.get().strip().upper() or None
        db_commit(
            "UPDATE magazzino_gomme SET marca=%s,modello=%s,misura=%s,quantita=%s,"
            "posizione=%s,stagione=%s,stato=%s,targa_veicolo=%s,"
            "deposito_pagato=%s,montate=%s,note=%s WHERE id_gomma=%s",
            (marca,
             e_modello.get().strip() or None,
             e_misura.get().strip() or None,
             qty,
             e_pos.get().strip() or None,
             stag_m.get(), stato_m.get(), targa_m,
             1 if dep_m.get() else 0,
             1 if mont_m.get() else 0,
             e_note_m.get("1.0","end").strip() or None,
             gid))
        messagebox.showinfo("OK","Aggiornato!",parent=win)
        win.destroy(); carica_gomme()

    tk.Button(win, text="💾 Salva modifiche", command=salva_m, width=20,
              bg="#4CAF50", fg="white", font=("Arial",10,"bold")).grid(
              row=6, column=0, columnspan=4, pady=14)

def on_doppio_clic_gomma(event=None):
    """
    Doppio clic intelligente:
    - colonna 'Dep. Pagato'  → alterna deposito_pagato
    - qualsiasi altra colonna → alterna montate
    """
    if not event: return
    iid = tree_gom.identify_row(event.y)
    if not iid or iid.startswith("sep_"): return
    try:
        gid = int(iid)
    except ValueError:
        return

    # Scopri su quale colonna ha cliccato
    col_id  = tree_gom.identify_column(event.x)   # es. "#10"
    try:
        col_idx = int(col_id.replace("#","")) - 1   # 0-based
    except ValueError:
        col_idx = -1

    # cols_gom = ("montate","cliente","targa_vei","marca","modello","stagione","misura","qty","stato","posizione","dep")
    # indice 10 = "dep"
    COL_DEP = 10

    rows = db_fetch("SELECT montate, deposito_pagato FROM magazzino_gomme WHERE id_gomma=%s",(gid,))
    if not rows: return
    r = rows[0]

    if col_idx == COL_DEP:
        # Alterna deposito pagato
        nuovo = 0 if bool(r["deposito_pagato"]) else 1
        db_commit("UPDATE magazzino_gomme SET deposito_pagato=%s WHERE id_gomma=%s",(nuovo, gid))
    else:
        # Alterna montate — UN SOLO set per cliente può essere montato
        nuovo = 0 if bool(r["montate"]) else 1
        if nuovo == 1:
            # Togli montate a tutte le altre gomme dello stesso cliente e stessa targa
            rows_cl = db_fetch("SELECT id_cliente, targa_veicolo FROM magazzino_gomme WHERE id_gomma=%s",(gid,))
            if rows_cl:
                cid_g    = rows_cl[0]["id_cliente"]
                targa_g  = rows_cl[0]["targa_veicolo"]
                if cid_g is not None:
                    # togli montate a tutti i set dello stesso cliente (stessa targa o qualsiasi)
                    db_commit(
                        "UPDATE magazzino_gomme SET montate=0 WHERE id_cliente=%s AND id_gomma!=%s",
                        (cid_g, gid))
        db_commit("UPDATE magazzino_gomme SET montate=%s WHERE id_gomma=%s",(nuovo, gid))

    carica_gomme()
    try: tree_gom.selection_set(str(gid)); tree_gom.see(str(gid))
    except: pass

tree_gom.bind("<Double-1>", on_doppio_clic_gomma)

def elimina_gomma():
    sel = tree_gom.selection()
    if not sel:
        messagebox.showwarning("Attenzione","Seleziona una riga"); return
    if sel[0].startswith("sep_"):
        messagebox.showwarning("Attenzione","Seleziona una riga gomma, non il separatore"); return
    if not messagebox.askyesno("Conferma","Eliminare questa voce dal magazzino?"): return
    db_commit("DELETE FROM magazzino_gomme WHERE id_gomma=%s",(int(sel[0]),)); carica_gomme()

btn_gom = tk.Frame(tab_gomme); btn_gom.pack(fill="x",padx=10,pady=9)
tk.Button(btn_gom,text="➕ Aggiungi Gomme",command=finestra_aggiungi_gomma,width=30,font=("Arial",11,"bold")).pack(side="left",padx=10)
tk.Button(btn_gom,text="✏️ Modifica",       command=modifica_gomma,         width=30,font=("Arial",11,"bold")).pack(side="left",padx=10)
tk.Button(btn_gom,text="🗑 Elimina",        command=elimina_gomma,          width=30,font=("Arial",11,"bold")).pack(side="left",padx=10)


# ══════════════════════════════════════════════════════════════
# SCHEDA 5 – STORICO FATTURE
# ══════════════════════════════════════════════════════════════
tab_storico_fat = tk.Frame(nb)
nb.add(tab_storico_fat, text="📂  Storico Fatture")

filtro_fat_frame = tk.Frame(tab_storico_fat); filtro_fat_frame.pack(fill="x",padx=10,pady=6)
tk.Label(filtro_fat_frame,text="🔍 Cerca cliente:").pack(side="left",padx=4)
filtro_fat_var = tk.StringVar()
filtro_fat_combo = ttk.Combobox(filtro_fat_frame, textvariable=filtro_fat_var, width=26)
filtro_fat_combo.pack(side="left",padx=4)
tk.Label(filtro_fat_frame, text="(scrivi per filtrare)", fg="gray", font=("Arial",8)).pack(side="left",padx=(0,6))
tk.Button(filtro_fat_frame,text="Mostra Tutte",command=lambda:[filtro_fat_var.set(""),carica_storico_fatture()]).pack(side="left",padx=4)

filtro_fat_clienti_tutti = []   # lista completa nomi clienti per il suggerimento

def _fat_aggiorna_suggerimenti(*_):
    """Filtra i suggerimenti e aggiorna la tabella in tempo reale."""
    testo = filtro_fat_var.get().lower()
    if testo:
        filtrati = [k for k in filtro_fat_clienti_tutti if testo in k.lower()]
    else:
        filtrati = filtro_fat_clienti_tutti
    filtro_fat_combo["values"] = filtrati
    carica_storico_fatture()

filtro_fat_var.trace_add("write", _fat_aggiorna_suggerimenti)
filtro_fat_combo.bind("<<ComboboxSelected>>", lambda e: carica_storico_fatture())

lf_sf = tk.LabelFrame(tab_storico_fat,text="Fatture  (doppio clic per aprire l'Excel)")
lf_sf.pack(fill="both",expand=True,padx=10,pady=4)
cols_sf = ("numero","data","cliente","veicolo","totale","pagata")
tree_sf = ttk.Treeview(lf_sf,columns=cols_sf,show="headings")
for c,w,h in zip(cols_sf,(80,105,200,210,115,100),("N°","Data","Cliente","Veicolo","Totale CHF","Stato")):
    tree_sf.heading(c,text=h); tree_sf.column(c,width=w)
sc_sf = ttk.Scrollbar(lf_sf,orient="vertical",command=tree_sf.yview)
tree_sf.config(yscrollcommand=sc_sf.set)
tree_sf.pack(side="left",fill="both",expand=True,padx=4,pady=4)
sc_sf.pack(side="right",fill="y",pady=4)
tree_sf.tag_configure("pagata",    background="#e8f5e9")
tree_sf.tag_configure("nonpagata", background="#fff3e0")

def apri_fattura_da_iid(iid):
    rows = db_fetch("SELECT file_path FROM fatture WHERE id_fattura=%s",(int(iid),))
    path = rows[0]["file_path"] if rows else None
    if not path:
        messagebox.showinfo("Info","Nessun file Excel associato a questa fattura"); return
    full = os.path.join(BASE_DIR, path)
    if not os.path.exists(full):
        messagebox.showwarning("Attenzione",f"File non trovato:\n{full}"); return
    apri_file(full)

def on_doppio_clic_fattura(event=None):
    iid = tree_sf.identify_row(event.y) if event else None
    if iid: apri_fattura_da_iid(iid)

tree_sf.bind("<Double-1>", on_doppio_clic_fattura)

def carica_storico_fatture():
    global filtro_fat_clienti_tutti
    tree_sf.delete(*tree_sf.get_children())
    filtro = f"%{filtro_fat_var.get()}%"

    # Aggiorna la lista suggerimenti clienti (solo se vuota o cambio scheda)
    if not filtro_fat_clienti_tutti:
        rows_cli = db_fetch("SELECT DISTINCT CONCAT(c.cognome,' ',c.nome) AS n FROM fatture f JOIN clienti c ON c.id_cliente=f.id_cliente ORDER BY n")
        filtro_fat_clienti_tutti = [r["n"] for r in rows_cli]
        testo = filtro_fat_var.get().lower()
        filtro_fat_combo["values"] = [k for k in filtro_fat_clienti_tutti if testo in k.lower()] if testo else filtro_fat_clienti_tutti

    totale_pagate = 0.0
    for r in db_fetch("""
        SELECT f.id_fattura,f.data_fattura,f.totale,f.pagata,
               CONCAT(c.cognome,' ',c.nome) AS cliente,
               CONCAT(IFNULL(v.targa,''),' ',IFNULL(v.marca,''),' ',IFNULL(v.modello,'')) AS veicolo
        FROM fatture f
        JOIN clienti c ON c.id_cliente=f.id_cliente
        JOIN veicoli v ON v.id_veicolo=f.id_veicolo
        WHERE CONCAT(c.cognome,' ',c.nome) LIKE %s
        ORDER BY f.id_fattura DESC
    """,(filtro,)):
        pag  = "✅ Pagata" if r["pagata"] else "❌ Non pagata"
        tag  = "pagata" if r["pagata"] else "nonpagata"
        tree_sf.insert("","end",iid=str(r["id_fattura"]),tags=(tag,),
            values=(r["id_fattura"],str(r["data_fattura"]) if r["data_fattura"] else "",
                    r["cliente"],r["veicolo"].strip(),
                    f"CHF {r['totale']:.2f}" if r["totale"] else "CHF 0.00",pag))
        if r["pagata"] and r["totale"]:
            totale_pagate += float(r["totale"])
    try:
        lbl_totale_sf.config(text=f"Totale pagate (IVA incl.): CHF {totale_pagate:.2f}")
    except Exception:
        pass

def rileggi_totale_da_excel():
    """Rilegge H45 dal file Excel e aggiorna il totale nel DB."""
    sel = tree_sf.selection()
    if not sel:
        messagebox.showwarning("Attenzione","Seleziona una fattura"); return
    fid = int(sel[0])
    rows = db_fetch("SELECT file_path, id_storico, totale FROM fatture WHERE id_fattura=%s",(fid,))
    if not rows:
        messagebox.showwarning("Errore","Fattura non trovata"); return
    r = rows[0]
    if not r["file_path"]:
        messagebox.showwarning("Errore","Nessun file Excel associato"); return
    path = os.path.join(BASE_DIR, r["file_path"])
    if not os.path.exists(path):
        messagebox.showwarning("Errore",f"File non trovato:\n{path}"); return

    totale = None

    # Prova win32com prima
    try:
        import win32com.client
        nome_file = os.path.basename(path)
        xl = None
        try:
            xl = win32com.client.GetActiveObject("Excel.Application")
            for wb in xl.Workbooks:
                if os.path.basename(wb.FullName).lower() == nome_file.lower():
                    ws = wb.Sheets("Ricevuta")
                    totale = ws.Range("H46").Value
                    break
        except Exception:
            pass

        if not totale:
            xl2 = win32com.client.Dispatch("Excel.Application")
            xl2.Visible = False
            xl2.DisplayAlerts = False
            wb2 = xl2.Workbooks.Open(path, ReadOnly=True, UpdateLinks=False)
            wb2.Application.CalculateFull()
            ws2 = wb2.Sheets("Ricevuta")
            totale = ws2.Range("H46").Value
            wb2.Close(False)
            xl2.Quit()
    except Exception:
        pass

    # Fallback openpyxl
    if not totale:
        try:
            from openpyxl import load_workbook
            wb3 = load_workbook(path, data_only=True)
            ws3 = wb3["Ricevuta"]
            totale = ws3["H46"].value
            if not totale:
                # somma manuale + IVA
                imp = sum(
                    float(ws3[f"E{row}"].value or 0) * float(ws3[f"F{row}"].value or 0)
                    for row in range(18, 44) if ws3[f"B{row}"].value)
                totale = round(imp * (1 + IVA), 2) if imp > 0 else None
        except Exception:
            pass

    if not totale:
        messagebox.showerror("Errore","Impossibile leggere il totale dal file Excel.\nProva ad aprire il file prima."); return

    totale = round(float(totale), 2)
    db_commit("UPDATE fatture SET totale=%s, iva_applicata=TRUE WHERE id_fattura=%s",(totale, fid))
    if r["id_storico"]:
        db_commit("UPDATE storico_interventi SET costo=%s WHERE id_intervento=%s",(totale, r["id_storico"]))
    carica_storico_fatture()
    messagebox.showinfo("✅ Aggiornato",f"Totale aggiornato: CHF {totale:.2f}")
    sel = tree_sf.selection()
    if not sel:
        messagebox.showwarning("Attenzione","Seleziona una fattura"); return
    db_commit("UPDATE fatture SET pagata=TRUE WHERE id_fattura=%s",(int(sel[0]),))
    carica_storico_fatture()

def segna_pagata():
    sel = tree_sf.selection()
    if not sel:
        messagebox.showwarning("Attenzione","Seleziona una fattura"); return
    db_commit("UPDATE fatture SET pagata=TRUE WHERE id_fattura=%s",(int(sel[0]),))
    carica_storico_fatture()

def togli_pagata():
    sel = tree_sf.selection()
    if not sel:
        messagebox.showwarning("Attenzione","Seleziona una fattura"); return
    fid = int(sel[0])
    vals = tree_sf.item(sel[0])["values"]
    if not messagebox.askyesno("Conferma",
        f"Rimuovere il pagamento dalla fattura N° {vals[0]} – {vals[2]}?\n\n"
        "La fattura tornerà a 'Non pagata'."): return
    db_commit("UPDATE fatture SET pagata=FALSE WHERE id_fattura=%s",(fid,))
    carica_storico_fatture()

def apri_fattura():
    sel = tree_sf.selection()
    if not sel:
        messagebox.showwarning("Attenzione","Seleziona una fattura"); return
    apri_fattura_da_iid(sel[0])

def elimina_fattura():
    sel = tree_sf.selection()
    if not sel:
        messagebox.showwarning("Attenzione","Seleziona una fattura"); return
    fid  = int(sel[0])
    vals = tree_sf.item(sel[0])["values"]

    # Finestra di conferma con checkbox storico
    win_conf = tk.Toplevel(root)
    win_conf.title("Conferma eliminazione")
    win_conf.geometry("420x200")
    win_conf.grab_set(); win_conf.resizable(False, False)

    tk.Label(win_conf,
        text=f"Eliminare la fattura N° {vals[0]} del {vals[1]}\nclientе: {vals[2]}?",
        font=("Arial",10), justify="center").pack(pady=(16,8), padx=16)

    elimina_storico_var = tk.BooleanVar(value=True)
    tk.Checkbutton(win_conf,
        text="Elimina anche lo storico veicolo associato",
        variable=elimina_storico_var,
        font=("Arial",9)).pack(padx=16, anchor="w")

    def conferma():
        win_conf.destroy()
        rows = db_fetch(
            "SELECT file_path, id_storico FROM fatture WHERE id_fattura=%s",(fid,))
        if rows:
            # Elimina file Excel
            if rows[0]["file_path"]:
                try:
                    percorso = os.path.join(BASE_DIR, rows[0]["file_path"])
                    if os.path.exists(percorso):
                        os.remove(percorso)
                except Exception:
                    pass
            # Elimina storico se spuntato
            if elimina_storico_var.get() and rows[0]["id_storico"]:
                db_commit("DELETE FROM storico_interventi WHERE id_intervento=%s",
                          (rows[0]["id_storico"],))

        db_commit("DELETE FROM fatture WHERE id_fattura=%s",(fid,))
        carica_storico_fatture()
        try: aggiorna_lbl_num_fat()
        except Exception: pass

    btn_row = tk.Frame(win_conf); btn_row.pack(pady=16)
    tk.Button(btn_row, text="🗑 Elimina", command=conferma,
              bg="#e53935", fg="white", font=("Arial",10,"bold"), width=14).pack(side="left", padx=8)
    tk.Button(btn_row, text="Annulla", command=win_conf.destroy,
              width=10).pack(side="left", padx=8)

btn_sf = tk.Frame(tab_storico_fat); btn_sf.pack(fill="x",padx=10,pady=9)
tk.Button(btn_sf,text="✅ Segna Pagata",    command=segna_pagata,          width=20,font=("Arial",11,"bold")).pack(side="left",padx=4)
tk.Button(btn_sf,text="↩️ Togli Pagamento", command=togli_pagata,           width=20,font=("Arial",11,"bold")).pack(side="left",padx=4)
tk.Button(btn_sf,text="📂 Apri Excel",      command=apri_fattura,           width=20,font=("Arial",11,"bold"),bg="#53d11c").pack(side="left",padx=4)
tk.Button(btn_sf,text="🔄 Rileggi Totale",  command=rileggi_totale_da_excel,width=20,font=("Arial",11,"bold"),bg="#2196F3",fg="white").pack(side="left",padx=4)
tk.Button(btn_sf,text="🗑 Elimina",         command=elimina_fattura,         width=20,font=("Arial",11,"bold")).pack(side="left",padx=4)
tk.Button(btn_sf,text="🔄 Aggiorna",        command=carica_storico_fatture,  width=20,font=("Arial",11,"bold")).pack(side="left",padx=4)
lbl_totale_sf = tk.Label(btn_sf, text="Totale pagate: CHF 0.00",
                          font=("Arial",10,"bold"), fg="#1a7a16")
lbl_totale_sf.pack(side="right", padx=10)


# ══════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════
# SINCRONIZZAZIONE ONEDRIVE
# ══════════════════════════════════════════════════════════════
def _trova_onedrive():
    """Trova la cartella OneDrive dell'utente corrente."""
    import winreg
    try:
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
            r"Software\Microsoft\OneDrive", 0, winreg.KEY_READ)
        path, _ = winreg.QueryValueEx(key, "UserFolder")
        winreg.CloseKey(key)
        if path and os.path.isdir(path):
            return path
    except Exception:
        pass
    # Fallback: cerca manualmente
    for candidato in [
        os.path.expandvars(r"%USERPROFILE%\OneDrive"),
        os.path.expandvars(r"%USERPROFILE%\OneDrive - Personal"),
        os.path.expandvars(r"%OneDriveConsumer%"),
        os.path.expandvars(r"%OneDrive%"),
    ]:
        if os.path.isdir(candidato):
            return candidato
    return None

ONEDRIVE_SYNC_DIR = None
ONEDRIVE_SQL_FILE = None

def _init_onedrive_path():
    global ONEDRIVE_SYNC_DIR, ONEDRIVE_SQL_FILE
    # Il file sync è direttamente nella cartella del gestionale (già su OneDrive)
    ONEDRIVE_SYNC_DIR = BASE_DIR
    ONEDRIVE_SQL_FILE = os.path.join(BASE_DIR, "officina_sync.sql")

def _timestamp_sql(path):
    """Restituisce la data/ora salvata nella prima riga del file SQL."""
    try:
        with open(path, "r", encoding="utf-8") as f:
            prima = f.readline()
        # formato: -- SYNC: 2026-04-25T14:30:00
        if "-- SYNC:" in prima:
            return datetime.datetime.fromisoformat(prima.split("SYNC:")[1].strip())
    except Exception:
        pass
    return None

def _trova_mysql_bin():
    """Trova il percorso di mysql.exe e mysqldump.exe cercando in Program Files."""
    import glob
    candidati = glob.glob(r"C:\Program Files\MySQL\MySQL Server *\bin")
    candidati += glob.glob(r"C:\Program Files (x86)\MySQL\MySQL Server *\bin")
    candidati += [r"C:\xampp\mysql\bin", r"C:\wamp64\bin\mysql\mysql8.0\bin"]
    for cartella in sorted(candidati, reverse=True):
        dump = os.path.join(cartella, "mysqldump.exe")
        mysql = os.path.join(cartella, "mysql.exe")
        if os.path.exists(dump) and os.path.exists(mysql):
            return dump, mysql
    # Prova nel PATH
    return "mysqldump", "mysql"

def sincronizza_onedrive_export(silenzioso=True):
    """Copia officina.db su OneDrive per sincronizzazione."""
    if not ONEDRIVE_SYNC_DIR:
        if not silenzioso:
            messagebox.showwarning("Sync", "OneDrive non trovato su questo PC.")
        return
    try:
        db_src = os.path.join(BASE_DIR, "officina.db")
        if not os.path.exists(db_src):
            if not silenzioso:
                messagebox.showwarning("Sync", "Database officina.db non trovato.")
            return
        os.makedirs(ONEDRIVE_SYNC_DIR, exist_ok=True)
        db_dst = os.path.join(ONEDRIVE_SYNC_DIR, "officina_sync.db")
        # Scrivi timestamp nel file di metadati
        ora = datetime.datetime.now().isoformat(timespec="seconds")
        with open(ONEDRIVE_SQL_FILE, "w", encoding="utf-8") as f:
            f.write(f"-- SYNC: {ora}\n")
        shutil.copy2(db_src, db_dst)
        if not silenzioso:
            messagebox.showinfo("✅ Sync OneDrive",
                f"Database sincronizzato su OneDrive!\n{db_dst}")
    except Exception as e:
        if not silenzioso:
            messagebox.showerror("Errore Sync", str(e))

def sincronizza_onedrive_import():
    """
    All'avvio: se il file OneDrive è più recente del DB locale, chiede
    se importare. Copia officina_sync.db → officina.db.
    """
    if not ONEDRIVE_SYNC_DIR:
        return
    db_remote = os.path.join(ONEDRIVE_SYNC_DIR, "officina_sync.db")
    if not os.path.exists(ONEDRIVE_SQL_FILE) or not os.path.exists(db_remote):
        return
    try:
        ts_remote = _timestamp_sql(ONEDRIVE_SQL_FILE)
        if not ts_remote:
            return
        ts_log_path = os.path.join(BASE_DIR, "ultimo_sync.txt")
        ts_local = None
        if os.path.exists(ts_log_path):
            try:
                with open(ts_log_path) as f:
                    ts_local = datetime.datetime.fromisoformat(f.read().strip())
            except Exception:
                pass
        if ts_local and ts_remote <= ts_local:
            return

        diff = ts_remote - (ts_local or datetime.datetime(2000,1,1))
        minuti = int(diff.total_seconds() / 60)
        if minuti < 60:      diff_str = f"{minuti} minuti fa"
        elif minuti < 1440:  diff_str = f"{minuti//60} ore fa"
        else:                diff_str = f"{minuti//1440} giorni fa"

        risposta = messagebox.askyesno(
            "📥 Dati più recenti su OneDrive",
            f"Su OneDrive c'è un backup più recente del database:\n\n"
            f"  📅 Salvato: {ts_remote.strftime('%d/%m/%Y %H:%M')}  ({diff_str})\n\n"
            f"Vuoi importarlo? (sovrascriverà i dati locali)\n\n"
            f"Premi NO per continuare con i dati locali.")
        if risposta:
            db_local = os.path.join(BASE_DIR, "officina.db")
            # Backup del db locale prima di sovrascrivere
            bak = db_local.replace(".db", "_pre_sync.db")
            if os.path.exists(db_local):
                shutil.copy2(db_local, bak)
            shutil.copy2(db_remote, db_local)
            with open(ts_log_path, "w") as f:
                f.write(ts_remote.isoformat())
            messagebox.showinfo("✅ Importato",
                "Dati aggiornati da OneDrive!\nRicarico il gestionale...")
            aggiorna_lista_clienti()
            carica_fat_clienti()
            carica_sto_clienti()
            carica_gom_clienti(); carica_gomme()
    except Exception:
        pass

def on_chiusura():
    """Esporta su OneDrive prima di chiudere."""
    threading.Thread(target=sincronizza_onedrive_export, args=(True,), daemon=False).start()
    import time; time.sleep(1)   # aspetta che finisca
    root.destroy()

# BACKUP AUTOMATICO SETTIMANALE SU CHIAVETTA
# ══════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════
# SISTEMA AGGIORNAMENTI AUTOMATICI
# ══════════════════════════════════════════════════════════════

def _confronta_versioni(v1, v2):
    """Restituisce True se v2 > v1 (es. '1.0.1' > '1.0.0')"""
    try:
        t1 = tuple(int(x) for x in v1.strip().split("."))
        t2 = tuple(int(x) for x in v2.strip().split("."))
        return t2 > t1
    except Exception:
        return False

def controlla_aggiornamenti(silenzioso=True):
    """
    Controlla se c'è una versione più recente su GitHub.
    Se silenzioso=True non mostra niente se non c'è aggiornamento.
    """
    if GITHUB_USER == "TUO_USERNAME_GITHUB":
        if not silenzioso:
            messagebox.showwarning("Aggiornamenti",
                "GitHub non configurato.\n\n"
                "Modifica GITHUB_USER e GITHUB_REPO in cima al file gui.py.")
        return
    try:
        import urllib.request
        with urllib.request.urlopen(GITHUB_VERSION_URL, timeout=5) as r:
            versione_remota = r.read().decode().strip()

        if _confronta_versioni(VERSIONE_CORRENTE, versione_remota):
            # Esegui nel thread principale (siamo già in un thread)
            root.after(0, lambda: _mostra_popup_aggiornamento(versione_remota))
        else:
            if not silenzioso:
                messagebox.showinfo("Aggiornamenti",
                    f"Il gestionale è aggiornato.\nVersione corrente: {VERSIONE_CORRENTE}")
    except Exception as ex:
        if not silenzioso:
            messagebox.showwarning("Aggiornamenti",
                f"Impossibile controllare gli aggiornamenti.\n"
                f"Verifica la connessione internet.\n\n{ex}")

def _mostra_popup_aggiornamento(versione_nuova):
    """Popup che chiede se aggiornare."""
    win = tk.Toplevel(root)
    win.title("🆕 Aggiornamento disponibile")
    win.geometry("420x220")
    win.grab_set(); win.resizable(False, False)

    tk.Label(win,
        text="🆕 Nuovo aggiornamento disponibile!",
        font=("Arial",13,"bold"), fg="#1565C0").pack(pady=(20,4))
    tk.Label(win,
        text=f"Versione corrente:  {VERSIONE_CORRENTE}\n"
             f"Versione disponibile:  {versione_nuova}",
        font=("Arial",11), justify="center").pack(pady=4)
    tk.Label(win,
        text="Vuoi aggiornare ora?\n"
             "Il gestionale si riavvierà automaticamente.",
        font=("Arial",10), fg="#555", justify="center").pack(pady=4)

    def esegui_aggiornamento():
        win.destroy()
        _scarica_e_installa()

    btn_row = tk.Frame(win); btn_row.pack(pady=14)
    tk.Button(btn_row, text="⬇️ Aggiorna ora", command=esegui_aggiornamento,
              bg="#1565C0", fg="white", font=("Arial",11,"bold"), width=16).pack(side="left", padx=8)
    tk.Button(btn_row, text="Dopo", command=win.destroy,
              width=10).pack(side="left", padx=8)

def _scarica_e_installa():
    """Scarica il nuovo gui.py e riavvia il programma."""
    try:
        import urllib.request

        # Finestra di progresso
        win_prog = tk.Toplevel(root)
        win_prog.title("Download aggiornamento...")
        win_prog.geometry("340x100")
        win_prog.grab_set(); win_prog.resizable(False, False)
        tk.Label(win_prog, text="⬇️ Download in corso...",
                 font=("Arial",11)).pack(pady=20)
        win_prog.update()

        # Scarica il nuovo gui.py
        percorso_gui = os.path.join(BASE_DIR, "gui.py")
        percorso_bak = os.path.join(BASE_DIR, "gui_backup.py")

        # Backup del vecchio
        shutil.copy2(percorso_gui, percorso_bak)

        with urllib.request.urlopen(GITHUB_GUI_URL, timeout=30) as r:
            nuovo_codice = r.read()

        with open(percorso_gui, "wb") as f:
            f.write(nuovo_codice)

        win_prog.destroy()
        messagebox.showinfo("✅ Aggiornamento completato",
            "Il gestionale è stato aggiornato!\n\n"
            "Si riavvierà ora automaticamente.")

        # Riavvia il processo
        python_exe = sys.executable
        subprocess.Popen([python_exe, percorso_gui])
        root.after(500, lambda: sys.exit(0))

    except Exception as ex:
        try: win_prog.destroy()
        except Exception: pass
        messagebox.showerror("Errore aggiornamento",
            f"Aggiornamento fallito:\n{ex}\n\n"
            "Il file precedente è stato mantenuto.")

BACKUP_LOG = os.path.join(BASE_DIR, "ultimo_backup.txt")

def trova_tutte_chiavette():
    """Restituisce lista di tuple (lettera, dimensione_gb) per ogni chiavetta USB."""
    chiavette = []
    try:
        import string, ctypes
        bitmask = ctypes.windll.kernel32.GetLogicalDrives()
        for lettera in string.ascii_uppercase:
            if bitmask & 1:
                drive = f"{lettera}:\\"
                if ctypes.windll.kernel32.GetDriveTypeW(drive) == 2:  # DRIVE_REMOVABLE
                    total = ctypes.c_ulonglong(0)
                    ctypes.windll.kernel32.GetDiskFreeSpaceExW(drive,None,ctypes.byref(total),None)
                    gb = round(total.value / (1024**3), 1)
                    chiavette.append((drive, gb))
            bitmask >>= 1
    except Exception:
        pass
    return chiavette

def trova_chiavetta():
    """Compatibilità: restituisce la prima chiavetta >50GB."""
    for drive, gb in trova_tutte_chiavette():
        if gb > 50:
            return drive
    return None

def _esegui_backup_su(chiavetta, silenzioso=True):
    """Esegue il backup sulla chiavetta specificata."""
    try:
        dest = os.path.join(chiavetta, "Backup_Garage_Tito",
                            f"backup_{datetime.date.today().isoformat()}")
        os.makedirs(dest, exist_ok=True)
        for item in os.listdir(BASE_DIR):
            src = os.path.join(BASE_DIR, item)
            dst = os.path.join(dest, item)
            try:
                if os.path.isdir(src):
                    if os.path.exists(dst): shutil.rmtree(dst)
                    shutil.copytree(src, dst)
                elif os.path.isfile(src):
                    shutil.copy2(src, dst)
            except Exception:
                pass
        # Backup SQLite — copia diretta del file .db
        db_src = os.path.join(BASE_DIR, "officina.db")
        if os.path.exists(db_src):
            shutil.copy2(db_src, os.path.join(dest, "officina.db"))
        oggi = datetime.date.today().isoformat()
        with open(BACKUP_LOG, "w") as f: f.write(oggi)
        try:
            root.after(0, lambda: lbl_backup.config(text=f"Ultimo backup: {oggi}"))
        except Exception:
            pass
        if not silenzioso:
            messagebox.showinfo("✅ Backup completato", f"Backup salvato su:\n{dest}")
    except Exception as e:
        if not silenzioso:
            messagebox.showerror("Errore Backup", str(e))

def esegui_backup(silenzioso=True):
    """Backup con selezione chiavetta se ce ne sono più di una."""
    chiavette = trova_tutte_chiavette()
    if not chiavette:
        if not silenzioso:
            messagebox.showwarning("Backup",
                "Nessuna chiavetta USB trovata.\nInseriscine una e riprova.")
        return

    if len(chiavette) == 1:
        # Una sola chiavetta: usa quella direttamente
        _esegui_backup_su(chiavette[0][0], silenzioso)
        return

    # Più chiavette: mostra finestra di selezione
    win_sel = tk.Toplevel(root)
    win_sel.title("Seleziona chiavetta per il backup")
    win_sel.geometry("360x220")
    win_sel.grab_set(); win_sel.resizable(False, False)

    tk.Label(win_sel, text="Scegli dove salvare il backup:",
             font=("Arial",11,"bold")).pack(pady=(16,8), padx=16)

    scelta_var = tk.StringVar(value=chiavette[0][0])
    for drive, gb in chiavette:
        tk.Radiobutton(win_sel,
            text=f"  {drive}  ({gb} GB)",
            variable=scelta_var, value=drive,
            font=("Arial",10)).pack(anchor="w", padx=32, pady=2)

    def avvia():
        drive_scelto = scelta_var.get()
        win_sel.destroy()
        threading.Thread(target=_esegui_backup_su,
                         args=(drive_scelto, False), daemon=True).start()

    btn_row = tk.Frame(win_sel); btn_row.pack(pady=14)
    tk.Button(btn_row, text="💾 Avvia Backup", command=avvia,
              bg="#1565C0", fg="white", font=("Arial",10,"bold"), width=16).pack(side="left", padx=8)
    tk.Button(btn_row, text="Annulla", command=win_sel.destroy, width=10).pack(side="left", padx=8)

def controlla_backup_settimanale():
    try:
        if os.path.exists(BACKUP_LOG):
            with open(BACKUP_LOG) as f:
                ultima = datetime.date.fromisoformat(f.read().strip())
            if (datetime.date.today() - ultima).days < 7:
                return
        # Backup silenzioso in background
        threading.Thread(target=esegui_backup, args=(True,), daemon=True).start()
    except Exception:
        pass

status_bar = tk.Frame(root,bd=1,relief="sunken",bg="#f0f0f0")
status_bar.pack(side="bottom",fill="x")
tk.Button(status_bar,text="🔄 Aggiornamenti",
          command=lambda: threading.Thread(
              target=lambda: controlla_aggiornamenti(False), daemon=True).start(),
          relief="flat",padx=8,bg="#e8f5e9").pack(side="right", padx=2)
tk.Button(status_bar,text="☁️ Sync OneDrive",
          command=lambda: threading.Thread(
              target=lambda: sincronizza_onedrive_export(False), daemon=True).start(),
          relief="flat",padx=8,bg="#e3f0ff").pack(side="right", padx=2)
tk.Button(status_bar,text="💾 Backup su Chiavetta",
          command=lambda: threading.Thread(target=esegui_backup,args=(False,),daemon=True).start(),
          relief="flat",padx=8,bg="#f0f0f0").pack(side="right")
ultimo = "mai"
try:
    if os.path.exists(BACKUP_LOG):
        with open(BACKUP_LOG) as f: ultimo = f.read().strip()
except: pass
lbl_backup = tk.Label(status_bar,text=f"Ultimo backup: {ultimo}",fg="gray",font=("Arial",9),bg="#f0f0f0")
lbl_backup.pack(side="right",padx=10)

# ── cambio scheda ─────────────────────────────────────────────
def on_tab_change(event=None):
    global filtro_fat_clienti_tutti
    tab = nb.index(nb.select())
    if tab==1:
        carica_fat_clienti()
        aggiorna_lbl_num_fat()
    elif tab==2: carica_sto_clienti()
    elif tab==3: carica_gom_clienti(); carica_gomme()
    elif tab==4:
        filtro_fat_clienti_tutti = []
        carica_storico_fatture()

nb.bind("<<NotebookTabChanged>>", on_tab_change)
aggiorna_lista_clienti()
controlla_backup_settimanale()
_riavvia_watcher_tutte_fatture()
_init_onedrive_path()
root.protocol("WM_DELETE_WINDOW", on_chiusura)
root.after(2000, sincronizza_onedrive_import)
# Controlla aggiornamenti in background 4 secondi dopo l'avvio
root.after(4000, lambda: threading.Thread(
    target=lambda: controlla_aggiornamenti(True), daemon=True).start())
root.mainloop()
